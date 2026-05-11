"""
╔══════════════════════════════════════════════════════════════════╗
║   TOKOPEDIA WILDLIFE SCRAPER — KLHK Edition                     ║
║   Versi  : 4.3.0                                                ║
║                                                                  ║
║   CHANGE LOG v4.3.0 (Threshold-Based Screenshot):              ║
║   ──────────────────────────────────────────────────            ║
║   [IMPROVE] Sistem validasi berbasis threshold field            ║
║          → Ekstrak SEMUA field dulu, hitung yang berhasil       ║
║          → Jika ≥ MIN_FIELDS_OK (default 5 dari 8) → langsung  ║
║            screenshot tanpa retry/reload sama sekali            ║
║          → Jika < MIN_FIELDS_OK → baru coba reload & retry      ║
║          → Jika semua retry habis tapi data parsial ada         ║
║            → tetap screenshot dengan status='partial'           ║
║          → Hanya skip screenshot jika 0 field berhasil          ║
║                                                                  ║
║   [BARU] Konstanta MIN_FIELDS_OK & TOTAL_FIELDS                 ║
║          → Mudah diubah di blok KONFIGURASI                     ║
║          → Default: 5/8 field = cukup untuk screenshot          ║
║                                                                  ║
║   [BARU] Status 'partial' di Excel                             ║
║          → Warna kuning muda — data ada tapi tidak lengkap      ║
║          → Berbeda dari 'no_data' (0 field) dan 'ok' (≥ 5/8)   ║
║                                                                  ║
║   CHANGE LOG v4.2.0 (Validasi Data + Infinite Loading):        ║
║   ──────────────────────────────────────────────────            ║
║   [FIX]  Screenshot hanya diambil SETELAH data valid            ║
║   [BARU] wait_for_product_data() — polling elemen kritis        ║
║   [BARU] detect_infinite_loading() — deteksi spinner hang       ║
║   [BARU] Kolom 'Status' di Excel                               ║
║                                                                  ║
║   CHANGE LOG v4.1.0 (Fix Tombol Muat Lebih Banyak):            ║
║   ──────────────────────────────────────────────────            ║
║   [FIX] Tombol diklik tapi produk tidak bertambah               ║
║   [FIX] Bug `or` pada prev_card_count                          ║
║   [FIX] Timing terlalu cepat setelah klik tombol               ║
║   [BARU] Deteksi throttle Tokopedia                             ║
║   [BARU] Fallback pagination URL (?page=N)                      ║
╚══════════════════════════════════════════════════════════════════╝
"""

import asyncio
import random
import re
import os
import json
import logging
import signal
import time
import sys
from datetime import datetime

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════
#  KONFIGURASI
# ══════════════════════════════════════════
OUTPUT_FOLDER       = "output"
KEYWORDS_FILE       = "keywords.txt"
PAGE_TIMEOUT        = 25
RESTART_EVERY       = 8
COOLDOWN_ON_HANG    = 15
MAX_LOAD_MORE       = 5        # 0 = klik semua tombol sampai habis
HARGA_THRESHOLD     = 350_000   # Rp — di atas ini → highlight merah
MAX_CONCURRENT_TABS = 5         # Maksimal tab terbuka bersamaan (ideal: 3-5)

# Validasi data sebelum screenshot
WAIT_DATA_TIMEOUT   = 15        # detik — batas tunggu elemen produk muncul
MAX_DETAIL_RETRY    = 2         # berapa kali coba reload sebelum skip
SPINNER_TIMEOUT     = 8         # detik — berapa lama spinner boleh hidup

# Threshold field — berapa field minimum yang harus berhasil diekstrak
# sebelum screenshot diambil. Total field yang dicek ada 8.
# Contoh: MIN_FIELDS_OK=5 → jika 5/8 atau lebih berhasil → langsung screenshot
#         Jika kurang dari 5 → coba reload dulu, baru screenshot setelah retry
#         Jika setelah retry masih kurang tapi > 0 → screenshot dengan status='partial'
#         Jika 0 field sama sekali → skip screenshot, status='no_data'
MIN_FIELDS_OK       = 5         # ubah ke angka lain sesuai kebutuhan (1–8)
TOTAL_FIELDS        = 8         # jumlah total field yang diekstrak (jangan diubah)

# Throttle detection
THROTTLE_LIMIT      = 3

CHECKPOINT_FILE     = "output/checkpoint.json"
LOG_FILE            = "output/scraper.log"
SCREENSHOT_FOLDER   = "output/screenshots"
SCREENSHOT_MAHAL    = "output/screenshots/mahal"
SUMMARY_FILE        = "output/ringkasan.xlsx"
PROXY_FILE          = "proxies.txt"

# [BARU] Toggle fitur — set False untuk disable
USE_PROXY           = True   # Proxy rotation dari proxies.txt
USE_ADAPTIVE_RL     = True   # Adaptive rate limiting
USE_RESOURCE_BLOCK  = True   # Block image/CSS di halaman search
USE_JSON_BACKUP     = True   # JSON Lines backup per keyword
USE_RETRY_QUEUE     = True   # Retry produk gagal di akhir sesi
USE_PROGRESS_DASH   = True   # Progress dashboard real-time
SCREENSHOT_JPEG     = True   # Simpan JPEG bukan PNG (hemat 80% disk)
SCREENSHOT_QUALITY  = 80     # Kualitas JPEG (1-100)


# ══════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════
def setup_logger() -> logging.Logger:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    logger = logging.getLogger("scraper")
    logger.setLevel(logging.INFO)
    if logger.handlers:
        return logger
    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger

logger = setup_logger()


# ══════════════════════════════════════════
#  CHECKPOINT
# ══════════════════════════════════════════
def checkpoint_load() -> dict:
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def checkpoint_save(data: dict):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def checkpoint_clear():
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)
        print("   🗑️  Checkpoint dihapus.")


# ══════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════
def load_keywords(path: str) -> list[str]:
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            f.write("gading gajah\ntaring harimau\nsisik trenggiling\n")
        print(f"⚠️  '{path}' tidak ditemukan. File contoh telah dibuat.")
        return ["gading gajah", "taring harimau", "sisik trenggiling"]
    with open(path, "r", encoding="utf-8") as f:
        kws = [line.strip() for line in f if line.strip()]
    if not kws:
        print("⚠️  keywords.txt kosong!")
        return []
    return kws

def clean_text(text: str) -> str:
    return re.sub(r'\s+', ' ', text).strip()

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name)

def parse_rating(raw: str) -> str:
    raw = clean_text(raw).replace(',', '.')
    match = re.search(r'\d+\.?\d*', raw)
    return match.group() if match else "N/A"

def extract_price_number(price_str: str) -> int:
    digits = re.sub(r'[^0-9]', '', str(price_str))
    return int(digits) if digits else 0

def format_price(raw: str) -> str:
    raw = raw.strip()
    if not raw:
        return "N/A"
    if not raw.startswith("Rp"):
        raw = "Rp" + raw
    return raw

def is_expensive(product: dict) -> bool:
    return extract_price_number(product.get("price", "0")) > HARGA_THRESHOLD

async def human_delay(min_sec=2.0, max_sec=5.0):
    if random.random() < 0.12:
        extra = random.uniform(5, 12)
        print(f"   💤 Jeda {extra:.1f}s...")
        await asyncio.sleep(extra)
    else:
        await asyncio.sleep(random.uniform(min_sec, max_sec))

async def scroll_and_extract(page, keyword: str, seen_links: set) -> list[dict]:
    js_code = """
    () => {
        return new Promise((resolve) => {
            function cleanText(text) {
                if (!text) return "N/A";
                return text.replace(/\\s+/g, ' ').trim();
            }

            let results = new Map();
            let lastHeight = document.body.scrollHeight;
            let noChangeCount = 0;
            
            let timer = setInterval(() => {
                window.scrollBy({top: 150, behavior: 'instant'});
                
                let cards = Array.from(document.querySelectorAll("div[data-testid='master-product-card']"));
                if (cards.length === 0) cards = Array.from(document.querySelectorAll("div.css-llwpbs"));
                if (cards.length === 0) cards = Array.from(document.querySelectorAll("div.css-5wh65g"));

                for (let card of cards) {
                    try {
                        let linkEl = card.querySelector("a[href*='tokopedia.com']");
                        let link = linkEl ? (linkEl.getAttribute("href") || "").split("?")[0] : "";
                        if (link && !results.has(link)) {
                            let titleEl  = card.querySelector("[data-testid='spnSRPProdName']");
                            let priceEl  = card.querySelector("[data-testid='spnSRPProdPrice']");
                            let ratingEl = card.querySelector("[data-testid='icnStarRating']");
                            let soldEl   = card.querySelector("[data-testid='txsImpSoldCount']");
                            let shopEl   = card.querySelector("[data-testid='lnkShopName']");

                            let title  = titleEl ? cleanText(titleEl.innerText) : "N/A";
                            let price  = priceEl ? cleanText(priceEl.innerText) : "";

                            let ratingRaw = ratingEl ? (ratingEl.getAttribute("aria-label") || "") : "";
                            let ratingMatch = ratingRaw.replace(/,/g, '.').match(/\\d+\\.?\\d*/);
                            let rating = ratingMatch ? ratingMatch[0] : "N/A";

                            let sold = soldEl ? cleanText(soldEl.innerText) : "N/A";
                            let shop = shopEl ? cleanText(shopEl.innerText) : "N/A";

                            results.set(link, {
                                title: title,
                                price: price,
                                rating: rating,
                                sold: sold,
                                shop: shop,
                                link: link
                            });
                        }
                    } catch (err) {}
                }

                let currentScroll = window.scrollY + window.innerHeight;
                let currentHeight = document.body.scrollHeight;
                
                if (currentScroll >= currentHeight - 100) {
                    if (currentHeight === lastHeight) {
                        noChangeCount++;
                        if (noChangeCount >= 10) { // 10 * 60ms = 600ms
                            clearInterval(timer);
                            resolve(Array.from(results.values()));
                        }
                    } else {
                        lastHeight = currentHeight;
                        noChangeCount = 0;
                    }
                } else {
                    noChangeCount = 0;
                }
            }, 60);
        });
    }
    """
    
    scraped_data = []
    try:
        scraped_data = await page.evaluate(js_code)
    except Exception as e:
        print(f"      [scroll/extract error: {e}]")

    products = []
    for item in scraped_data:
        link = item.get("link", "")
        if not link or link in seen_links:
            continue

        seen_links.add(link)
        item["keyword"]      = keyword
        item["price"]        = format_price(item.get("price", ""))
        item["description"]  = ""
        item["review_count"] = "N/A"
        item["stock"]        = "N/A"

        products.append(item)

    return products


# ══════════════════════════════════════════
#  [BARU] PROXY MANAGER
# ══════════════════════════════════════════
class ProxyManager:
    def __init__(self, path: str = PROXY_FILE):
        self.proxies: list[str] = []
        self.dead: set[str]    = set()
        self._idx: int         = 0
        if not USE_PROXY:
            return
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                self.proxies = [l.strip() for l in f if l.strip() and not l.startswith("#")]
            print(f"   🌐 Proxy dimuat: {len(self.proxies)} entri")
        else:
            open(path, "w").close()
            print(f"   ℹ️  {path} tidak ditemukan — buat file kosong, mode direct")

    def next_proxy(self) -> dict | None:
        if not self.proxies:
            return None
        available = [p for p in self.proxies if p not in self.dead]
        if not available:
            print("   ⚠️  Semua proxy mati — fallback direct")
            return None
        raw = available[self._idx % len(available)]
        self._idx += 1
        m = re.match(r'(https?|socks5?)://(?:([^:@]+):([^@]+)@)?([^:/]+):(\d+)', raw)
        if not m:
            return {"server": raw}
        proto, user, pw, host, port = m.groups()
        proxy: dict = {"server": f"{proto}://{host}:{port}"}
        if user:
            proxy["username"] = user
            proxy["password"] = pw
        return proxy

    def mark_dead(self, proxy: dict | None):
        if proxy:
            srv = proxy.get("server", "")
            if srv:
                self.dead.add(srv)
                print(f"   💀 Proxy mati: {srv}")


# ══════════════════════════════════════════
#  [BARU] ADAPTIVE RATE LIMITER
# ══════════════════════════════════════════
class AdaptiveRateLimit:
    def __init__(self, base_min: float = 1.0, base_max: float = 2.5):
        self.base_min      = base_min
        self.base_max      = base_max
        self.factor        = 1.0
        self._streak       = 0

    async def wait(self):
        if not USE_ADAPTIVE_RL:
            return
        delay = random.uniform(self.base_min, self.base_max) * self.factor
        await asyncio.sleep(delay)

    def on_success(self):
        self._streak += 1
        if self._streak >= 5:
            self.factor = max(0.5, self.factor * 0.75)
            self._streak = 0

    def on_failure(self):
        self.factor = min(8.0, self.factor * 2.0)
        self._streak = 0
        print(f"   ⚠️  Rate limit naik — delay factor={self.factor:.1f}x")


# ══════════════════════════════════════════
#  [BARU] PROGRESS TRACKER
# ══════════════════════════════════════════
class ProgressTracker:
    def __init__(self, total: int):
        self.total      = total
        self.done       = 0
        self.ok         = 0
        self.partial    = 0
        self.blocked    = 0
        self.no_data    = 0
        self._t0        = time.time()

    def update(self, status: str):
        if not USE_PROGRESS_DASH:
            return
        self.done += 1
        if   status == "ok":      self.ok      += 1
        elif status == "partial": self.partial  += 1
        elif status == "blocked": self.blocked  += 1
        else:                     self.no_data  += 1
        if self.done % 10 == 0 or self.done == self.total:
            elapsed = time.time() - self._t0
            speed   = (self.done / elapsed * 60) if elapsed > 0 else 0
            eta     = ((self.total - self.done) / (speed / 60)) if speed > 0 else 0
            pct     = self.done / self.total * 100
            print(
                f"\r   📊 [{self.done}/{self.total}] {pct:.1f}% | "
                f"⚡{speed:.1f}/min | ⏱️ETA {eta:.0f}m | "
                f"✅{self.ok} 🟡{self.partial} ⛔{self.blocked} ❌{self.no_data}   ",
                end="", flush=True
            )
            if self.done == self.total:
                print()


# ══════════════════════════════════════════
#  [BARU] JSON BACKUP (JSONL)
# ══════════════════════════════════════════
def _backup_path(keyword: str) -> str:
    return os.path.join(OUTPUT_FOLDER, f"backup_{sanitize_filename(keyword.lower())}.jsonl")

def backup_load_done_links(keyword: str) -> set:
    path = _backup_path(keyword)
    done: set = set()
    if not USE_JSON_BACKUP or not os.path.exists(path):
        return done
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            try:
                obj = json.loads(line)
                if obj.get("link"):
                    done.add(obj["link"])
            except Exception:
                pass
    return done

def backup_append(keyword: str, product: dict):
    if not USE_JSON_BACKUP:
        return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    try:
        with open(_backup_path(keyword), "a", encoding="utf-8") as f:
            f.write(json.dumps(product, ensure_ascii=False) + "\n")
    except Exception:
        pass


# ══════════════════════════════════════════
#  [BARU] GRACEFUL SHUTDOWN
# ══════════════════════════════════════════
_shutdown_requested = False

def _setup_signal_handler():
    def _handler(sig, frame):
        global _shutdown_requested
        _shutdown_requested = True
        print("\n\n   🛑 Ctrl+C — menyimpan data & keluar setelah batch ini selesai...")
    signal.signal(signal.SIGINT, _handler)
    try:
        signal.signal(signal.SIGTERM, _handler)
    except Exception:
        pass


def _unique_path(base: str) -> str:
    if not os.path.exists(base):
        return base
    root, ext = os.path.splitext(base)
    c = 1
    while os.path.exists(f"{root}_{c}{ext}"):
        c += 1
    return f"{root}_{c}{ext}"


# ══════════════════════════════════════════
#  SCREENSHOT
# ══════════════════════════════════════════
async def take_screenshot(page, product: dict, folder: str) -> str | None:
    try:
        os.makedirs(folder, exist_ok=True)
        keyword  = sanitize_filename(product.get("keyword", "unknown").lower())
        title    = sanitize_filename(product.get("title", "produk")[:40])
        price    = re.sub(r'[^0-9]', '', product.get("price", "0"))
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        # [MODIF] JPEG/WebP untuk hemat disk
        ext      = ".jpg" if SCREENSHOT_JPEG else ".png"
        filename = f"{ts}_{keyword}_{price}_{title}{ext}"
        filepath = os.path.join(folder, filename)
        if SCREENSHOT_JPEG:
            await page.screenshot(
                path=filepath, full_page=False,
                type="jpeg", quality=SCREENSHOT_QUALITY
            )
        else:
            await page.screenshot(path=filepath, full_page=True)
        print(f"      📷 Screenshot: {filepath}")
        logger.info(f"Screenshot: {filepath}")
        return filepath
    except Exception as e:
        print(f"      ⚠️  Screenshot gagal: {e}")
        return None


# ══════════════════════════════════════════
#  NOTIFIKASI WINDOWS
# ══════════════════════════════════════════
def notify_expensive(product: dict):
    title = product.get("title", "")[:60]
    price = product.get("price", "N/A")
    try:
        import winsound
        winsound.Beep(800, 400)
    except Exception:
        pass
    try:
        import subprocess
        safe_msg = f"Harga {price} | {title}".replace("'","").replace('"',"")
        ps = f"""
Add-Type -AssemblyName System.Windows.Forms
$n = New-Object System.Windows.Forms.NotifyIcon
$n.Icon = [System.Drawing.SystemIcons]::Warning
$n.Visible = $true
$n.ShowBalloonTip(6000, '💰 Produk Harga Tinggi!', '{safe_msg}', [System.Windows.Forms.ToolTipIcon]::Warning)
Start-Sleep 7
$n.Dispose()"""
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        subprocess.Popen(
            ["powershell", "-WindowStyle", "Hidden", "-Command", ps],
            creationflags=flags
        )
    except Exception:
        pass


# ══════════════════════════════════════════
#  BROWSER CONTEXT
# ══════════════════════════════════════════
BLOCK_PATTERNS = re.compile(
    r"(google-analytics|googletagmanager|doubleclick|facebook\.net"
    r"|fbcdn|gtm\.js|analytics|tracker|hotjar|mixpanel"
    r"|amplitude|segment\.io|newrelic|sentry\.io|ads\."
    r"|\.woff|\.woff2|\.ttf|\.otf|youtube\.com/embed"
    r"|player\.vimeo\.com|\.mp4|\.webm|tiktok\.com|js/yt)"
)
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
]
VIEWPORTS = [
    {"width": 1366, "height": 768},
    {"width": 1440, "height": 900},
    {"width": 1280, "height": 800},
    {"width": 1920, "height": 1080},
]

async def create_context(browser, proxy: dict = None):  # [MODIF] tambah proxy param
    ua = random.choice(USER_AGENTS)
    vp = random.choice(VIEWPORTS)
    proxy_str = proxy.get("server", "direct") if proxy else "direct"
    print(f"   🔄 Context baru — {vp['width']}x{vp['height']} | proxy={proxy_str}")
    ctx_opts = dict(
        user_agent=ua, viewport=vp,
        locale="id-ID", timezone_id="Asia/Jakarta",
        java_script_enabled=True, ignore_https_errors=True,
        extra_http_headers={
            "Accept"                   : "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language"          : "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept-Encoding"          : "gzip, deflate, br",
            "Connection"               : "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest"           : "document",
            "Sec-Fetch-Mode"           : "navigate",
            "Sec-Fetch-Site"           : "none",
            "Sec-Fetch-User"           : "?1",
            "Cache-Control"            : "max-age=0",
        },
    )
    if proxy:
        ctx_opts["proxy"] = proxy  # [BARU]
    context = await browser.new_context(**ctx_opts)

    async def block_trackers(route):
        if BLOCK_PATTERNS.search(route.request.url):
            await route.abort()
        else:
            await route.continue_()

    await context.route("**/*", block_trackers)
    await context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver',   { get: () => undefined });
        Object.defineProperty(navigator, 'plugins',     { get: () => [1, 2, 3, 4, 5] });
        Object.defineProperty(navigator, 'languages',   { get: () => ['id-ID', 'id', 'en-US'] });
        window.chrome = { runtime: {}, loadTimes: () => {}, csi: () => {}, app: {} };
        Object.defineProperty(navigator, 'permissions', {
            get: () => ({ query: (p) => Promise.resolve({
                state: p.name === 'notifications' ? 'denied' : 'granted'
            }) })
        });
        delete window.__playwright;
        delete window.__pw_manual;
    """)
    page = await context.new_page()
    page.set_default_timeout(PAGE_TIMEOUT * 1000)
    return context, page


# ══════════════════════════════════════════
#  NAVIGASI AMAN
# ══════════════════════════════════════════
async def is_page_hanging(page) -> bool:
    try:
        result = await asyncio.wait_for(
            page.evaluate("() => document.readyState"), timeout=5.0
        )
        return result == "loading"
    except Exception:
        return True

async def is_blocked(page) -> bool:
    try:
        title = (await page.title()).lower()
        url   = page.url.lower()
        if any(k in title for k in [
            "captcha","verify","robot","access denied","forbidden","blocked"
        ]):
            return True
        if any(k in url for k in ["cf-challenge","captcha","verify","blocked"]):
            return True
        for sel in ["iframe[src*='captcha']", "#captcha", ".g-recaptcha"]:
            if await page.query_selector(sel):
                return True
    except Exception:
        pass
    return False

async def safe_goto(page, url: str, label="halaman", retry=2) -> bool:
    for attempt in range(retry):
        print(f"   🌐 Membuka {label}{'  (retry)' if attempt > 0 else ''}...")
        try:
            # FIX: asyncio.wait_for timeout dalam detik, page.goto timeout dalam ms
            # Gunakan timeout langsung di page.goto agar konsisten
            await page.goto(
                url,
                wait_until="domcontentloaded",
                timeout=PAGE_TIMEOUT * 1000  # ms
            )
        except Exception as e:
            err_str = str(e).lower()
            if "timeout" in err_str or "time out" in err_str:
                print(f"   ⏰ Timeout — paksa stop...")
            else:
                print(f"   ❌ Error: {e}")
            try:
                await page.evaluate("window.stop()")
            except Exception:
                pass
            await asyncio.sleep(2)
            if attempt < retry - 1:
                continue
            return False

        await asyncio.sleep(random.uniform(1.5, 2.5))

        if await is_page_hanging(page):
            try:
                await page.evaluate("window.stop()")
            except Exception:
                pass
            await asyncio.sleep(1)
        return True
    return False


# ══════════════════════════════════════════
#  EXTRACT CARDS
# ══════════════════════════════════════════



# ══════════════════════════════════════════
#  KLIK TOMBOL MUAT LEBIH BANYAK
# ══════════════════════════════════════════
async def click_load_more(page) -> bool:
    await asyncio.sleep(1.0)

    try:
        btn_locator = page.locator("button", has_text="Muat Lebih Banyak").first
        count = await btn_locator.count()
        if count == 0:
            btn_locator = page.locator("button[data-unify='Button']").filter(
                has_text="Muat Lebih Banyak"
            ).first
            count = await btn_locator.count()
        if count == 0:
            return False
    except Exception:
        return False

    try:
        await btn_locator.scroll_into_view_if_needed()
        await asyncio.sleep(random.uniform(0.8, 1.3))

        box = await btn_locator.bounding_box()
        if box is None:
            return False

        tx = box["x"] + box["width"]  / 2 + random.uniform(-5, 5)
        ty = box["y"] + box["height"] / 2 + random.uniform(-3, 3)

        vp = page.viewport_size or {"width": 1280, "height": 800}
        start_x = random.uniform(vp["width"] * 0.1, vp["width"] * 0.4)
        start_y = random.uniform(vp["height"] * 0.3, vp["height"] * 0.6)
        await page.mouse.move(start_x, start_y, steps=8)
        await asyncio.sleep(random.uniform(0.1, 0.2))
        await page.mouse.move(tx, ty, steps=15)
        await asyncio.sleep(random.uniform(0.3, 0.6))

        await page.mouse.down()
        await asyncio.sleep(random.uniform(0.08, 0.18))
        await page.mouse.up()

        print(f"      Klik tombol 'Muat Lebih Banyak' di ({tx:.0f}, {ty:.0f})")

        # FIX: Hapus networkidle di sini — caller sudah punya waitnya sendiri
        # networkidle bisa block sampai 12 detik, tidak perlu dobel
        await asyncio.sleep(random.uniform(1.5, 2.5))

        return True

    except Exception as e:
        print(f"      Gagal klik: {e}")
        return False


async def count_cards_on_page(page) -> int:
    cards = await page.query_selector_all("div[data-testid='master-product-card']")
    if cards:
        return len(cards)
    cards = await page.query_selector_all("div.css-llwpbs")
    if cards:
        return len(cards)
    cards = await page.query_selector_all("div.css-5wh65g")
    return len(cards)


# ══════════════════════════════════════════
#  SCRAPE SEMUA PRODUK — LOAD MORE + FALLBACK PAGINATION
# ══════════════════════════════════════════
async def scrape_all_pages(
    page, keyword: str,
    browser=None, context_ref: list = None,
    is_cdp: bool = False
) -> tuple[list[dict], object, object]:

    all_products   = []
    seen_links     = set()
    load_count     = 0
    throttle_count = 0
    page_num       = 1
    use_pagination = False

    base_url = f"https://www.tokopedia.com/search?q={keyword.replace(' ', '+')}&navsource=home"

    print(f"\n   🌐 Membuka halaman search...")
    if not await safe_goto(page, base_url, label="search"):
        return all_products, context_ref[0] if context_ref else None, page

    if await is_blocked(page):
        print(f"   ⛔ Diblokir saat membuka search")
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)
        await page.screenshot(
            path=f"{OUTPUT_FOLDER}/blocked_{sanitize_filename(keyword)}.png"
        )
        return all_products, context_ref[0] if context_ref else None, page

    # [BARU] Block image/CSS di halaman search — hemat bandwidth & percepat load
    # Detail pages (new_page) tidak punya route ini, jadi screenshot tetap normal
    if USE_RESOURCE_BLOCK:
        _SEARCH_BLOCK = re.compile(
            r'\.(png|jpe?g|gif|webp|svg|ico|mp4|webm|css|woff2?|ttf|otf)(\?|$)',
            re.IGNORECASE
        )
        async def _block_search_res(route):
            if _SEARCH_BLOCK.search(route.request.url):
                await route.abort()
            else:
                await route.continue_()
        await page.route("**/*", _block_search_res)

    await asyncio.sleep(2)
    for sel in [
        "button[aria-label='close']",
        "button[class*='CloseButton']",
        "[data-testid='btnClosePromo']",
    ]:
        try:
            btn = await page.query_selector(sel)
            if btn and await btn.is_visible():
                await btn.click()
                await asyncio.sleep(1)
                break
        except Exception:
            pass

    await page.evaluate("window.scrollTo(0, 0)")
    await asyncio.sleep(1.0)

    while True:
        load_count += 1
        if MAX_LOAD_MORE > 0 and load_count > MAX_LOAD_MORE + 1:
            print(f"   🏁 Batas {MAX_LOAD_MORE} klik tercapai.")
            break

        print(f"\n   🔄 Putaran {load_count} — scroll & ambil produk...")

        before    = len(all_products)
        new_prods = await scroll_and_extract(page, keyword, seen_links)
        all_products.extend(new_prods)
        new_this  = len(all_products) - before

        print(f"      ✅ +{new_this} produk baru | Total: {len(all_products)}")
        logger.info(f"[{keyword}] putaran={load_count} baru={new_this} total={len(all_products)}")

        if not use_pagination:
            clicked = await click_load_more(page)

            if not clicked:
                if load_count == 1:
                    print(f"   ℹ️  Tidak ada tombol 'Muat lebih banyak' — fallback ke pagination URL")
                    use_pagination = True
                    page_num = 2
                    continue
                else:
                    print(f"   🏁 Tombol tidak ada lagi — semua produk sudah diambil.")
                    break

            print(f"      ⏳ Menunggu produk baru ter-load...")
            await asyncio.sleep(random.uniform(2.5, 4.0))

            print(f"      ⬇️  Melanjutkan scroll ke bawah...")
            await asyncio.sleep(1.5)

            if new_this == 0:
                throttle_count += 1
                print(f"      ⚠️  Tidak ada produk baru di putaran ini ({throttle_count}/{THROTTLE_LIMIT})")
                if throttle_count >= THROTTLE_LIMIT:
                    print(f"      🔄 Throttle limit tercapai — restart sesi")
                    if context_ref and browser:
                        if is_cdp:
                            try:
                                await page.close()
                            except Exception:
                                pass
                            await asyncio.sleep(random.uniform(5, 10))
                            page = await context_ref[0].new_page()
                        else:
                            try:
                                await context_ref[0].close()
                            except Exception:
                                pass
                            await asyncio.sleep(random.uniform(12, 20))
                            new_ctx, new_page = await create_context(browser)
                            context_ref[0] = new_ctx
                            page            = new_page

                        await safe_goto(page, base_url, label="search (setelah restart)")
                        await page.evaluate("window.scrollTo(0, 0)")
                        await asyncio.sleep(2.0)
                    throttle_count = 0
                    continue
            else:
                if throttle_count > 0:
                    print(f"      ✅ Produk kembali bertambah — reset throttle counter")
                throttle_count = 0

        else:
            next_url = (
                f"https://www.tokopedia.com/search"
                f"?q={keyword.replace(' ', '+')}&navsource=home&page={page_num}"
            )
            print(f"      ➡️  Buka halaman {page_num} via URL...")
            if not await safe_goto(page, next_url, label=f"hal.{page_num}"):
                break
            if await is_blocked(page):
                break

            await page.evaluate("window.scrollTo(0, 0)")
            await asyncio.sleep(1.5)

            cards_here = await count_cards_on_page(page)
            if cards_here == 0:
                print(f"      🏁 Halaman {page_num} kosong — selesai.")
                break

            if new_this == 0 and load_count > 1:
                throttle_count += 1
                if throttle_count >= THROTTLE_LIMIT:
                    break
            else:
                throttle_count = 0

            page_num += 1
            await asyncio.sleep(random.uniform(2.0, 4.0))

    current_context = context_ref[0] if context_ref else None
    print(f"\n   📦 Total: {len(all_products)} produk dari {load_count} putaran")
    return all_products, current_context, page


# ══════════════════════════════════════════
#  [BARU v4.2.0] TUNGGU DATA PRODUK
# ══════════════════════════════════════════
async def wait_for_product_data(page, timeout: int = WAIT_DATA_TIMEOUT) -> bool:
    """
    Polling elemen kritis produk sampai muncul.
    Return True  → data ada, aman untuk diekstrak + screenshot.
    Return False → timeout, halaman kosong / masih loading.
    """
    critical_selectors = [
        "h1[data-testid='lblPDPDetailProductName']",
        "div[data-testid='lblPDPDetailProductPrice']",
        "[data-testid='pdp_comp-product_content']",
    ]

    deadline = asyncio.get_event_loop().time() + timeout
    while asyncio.get_event_loop().time() < deadline:
        for sel in critical_selectors:
            try:
                el = await page.query_selector(sel)
                if el and await el.is_visible():
                    return True
            except Exception:
                pass
        await asyncio.sleep(0.8)

    return False


# ══════════════════════════════════════════
#  [BARU v4.2.0] DETEKSI INFINITE LOADING
# ══════════════════════════════════════════
async def detect_infinite_loading(page, timeout: int = SPINNER_TIMEOUT) -> bool:
    """
    Cek apakah spinner / skeleton masih tampil setelah N detik.
    Return True  → halaman hang (infinite loading).
    Return False → halaman normal atau tidak ada spinner.
    """
    spinner_selectors = [
        "[data-testid='skeleton-pdp']",
        ".skeleton-loading",
        "[class*='Skeleton']",
        "[class*='skeleton']",
        "[class*='shimmer']",
        "[class*='Shimmer']",
        "[class*='spinner']",
        "[class*='Spinner']",
        "[class*='loading']",
        "[role='progressbar']",
        "svg[class*='animate']",
    ]

    spinner_found = False
    for sel in spinner_selectors:
        try:
            el = await page.query_selector(sel)
            if el and await el.is_visible():
                spinner_found = True
                break
        except Exception:
            pass

    if not spinner_found:
        return False

    # Spinner ada — tunggu N detik, periksa apakah masih ada
    await asyncio.sleep(timeout)
    for sel in spinner_selectors:
        try:
            el = await page.query_selector(sel)
            if el and await el.is_visible():
                return True   # Masih ada → infinite loading
        except Exception:
            pass

    return False   # Sudah hilang → loading selesai normal


# ══════════════════════════════════════════
#  HELPER — Ekstrak semua field dari halaman detail
# ══════════════════════════════════════════
async def _extract_detail_fields(page, product: dict) -> int:
    """
    Coba ekstrak semua field dari halaman detail yang sudah terbuka.
    Return jumlah field yang berhasil diambil (0 – TOTAL_FIELDS).
    Tidak mengubah field yang sudah ada nilainya dari attempt sebelumnya
    kecuali nilai baru lebih baik (bukan N/A).
    """
    selectors = {
        "title"       : "h1[data-testid='lblPDPDetailProductName']",
        "price"       : "div[data-testid='lblPDPDetailProductPrice']",
        "rating"      : "span.main[data-testid='lblPDPDetailProductRatingNumber']",
        "review_count": "span[data-testid='lblPDPDetailProductRatingCounter']",
        "sold"        : "span[data-testid='lblPDPDetailProductSoldCounter']",
        "stock"       : "[data-testid='stock-label']",
        "description" : "[data-testid='lblPDPDescriptionProduk']",
        "shop"        : "[data-testid='llbPDPFooterShopName']",
    }

    extracted_count = 0
    for field, sel in selectors.items():
        try:
            el = await page.query_selector(sel)
            if el:
                raw = clean_text(await el.inner_text())
                if field == "price":
                    product[field] = format_price(raw)
                elif field == "rating":
                    product[field] = parse_rating(raw)
                elif field == "description":
                    product[field] = raw[:800] + "..." if len(raw) > 800 else raw
                elif field == "stock":
                    product[field] = re.sub(r'[^0-9.,]', '', raw) or raw
                else:
                    product[field] = raw
                extracted_count += 1
        except Exception:
            pass

    return extracted_count


# ══════════════════════════════════════════
#  SCRAPE DETAIL PRODUK + SCREENSHOT (v4.3.0)
# ══════════════════════════════════════════
async def scrape_product_detail(page, product: dict) -> dict:
    """
    Alur v4.3.0 — Threshold-Based Screenshot:

      1. Buka halaman detail
      2. Deteksi infinite loading → jika ya, paksa stop
      3. Tunggu minimal 1 elemen kritis muncul
      4. Ekstrak SEMUA field sekaligus → hitung yang berhasil

      5a. field ≥ MIN_FIELDS_OK  → langsung screenshot (tidak perlu retry)
      5b. field < MIN_FIELDS_OK  → reload & coba ekstrak lagi (max MAX_DETAIL_RETRY)
           - Setelah retry: field ≥ MIN_FIELDS_OK → screenshot, status='ok'
           - Setelah retry: 0 < field < MIN_FIELDS_OK → screenshot, status='partial'
           - Setelah retry: field = 0 → skip screenshot, status='no_data'
    """
    if not product.get("link"):
        product["scraped_at"]   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        product["status"]       = "no_link"
        product["screenshot"]   = ""
        product["harga_tinggi"] = "TIDAK"
        return product

    best_extracted = 0   # simpan hasil ekstraksi terbaik lintas attempt

    for attempt in range(1, MAX_DETAIL_RETRY + 1):
        label = f"detail (percobaan {attempt}/{MAX_DETAIL_RETRY})"

        # ── 1. Buka halaman ──────────────────────────────────────
        if not await safe_goto(page, product["link"], label=label):
            print(f"      ⚠️  Gagal buka halaman")
            await asyncio.sleep(3)
            continue

        if await is_blocked(page):
            print(f"      ⛔ Halaman di-block")
            product["scraped_at"]   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            product["status"]       = "blocked"
            product["screenshot"]   = ""
            product["harga_tinggi"] = "TIDAK"
            return product

        # ── 2. Deteksi infinite loading (fast check, 2s saja) ────
        # FIX: Kurangi SPINNER_TIMEOUT di sini jadi 2 detik saja
        # agar tidak memblokir tab-tab lain yang sedang paralel
        print(f"      🔍 Cek loading state...")
        is_hanging = await detect_infinite_loading(page, timeout=2)
        if is_hanging:
            print(f"      ♾️  Infinite loading — paksa stop...")
            try:
                await page.evaluate("window.stop()")
            except Exception:
                pass
            await asyncio.sleep(1.5)
            await page.evaluate("window.scrollTo(0, 300)")
            await asyncio.sleep(0.5)

        # ── 3. Tunggu minimal 1 elemen kritis ────────────────────
        print(f"      ⏳ Tunggu elemen produk ({WAIT_DATA_TIMEOUT}s maks)...")
        await page.evaluate("window.scrollTo(0, 400)")
        await asyncio.sleep(random.uniform(1.0, 1.8))
        await wait_for_product_data(page, timeout=WAIT_DATA_TIMEOUT)
        # Tidak langsung return jika False — biarkan ekstraksi menghitung sendiri

        # ── 4. Ekstrak semua field, hitung hasilnya ───────────────
        extracted = await _extract_detail_fields(page, product)
        best_extracted = max(best_extracted, extracted)

        print(
            f"      📋 {extracted}/{TOTAL_FIELDS} field diambil "
            f"(threshold: {MIN_FIELDS_OK}/{TOTAL_FIELDS})"
        )

        # ── 5a. Sudah cukup → langsung screenshot, tidak perlu retry
        if extracted >= MIN_FIELDS_OK:
            print(f"      ✅ Data cukup ({extracted}/{TOTAL_FIELDS}) — lanjut screenshot")
            product["status"] = "ok"
            break

        # ── 5b. Kurang dari threshold → coba reload jika masih ada attempt
        print(
            f"      ⚠️  Data kurang ({extracted}/{TOTAL_FIELDS} < {MIN_FIELDS_OK}) "
            f"— {'reload & coba lagi' if attempt < MAX_DETAIL_RETRY else 'retry habis'}"
        )
        logger.warning(
            f"LOW_FIELDS attempt={attempt} got={extracted}/{TOTAL_FIELDS} | "
            f"{product.get('link','')[:80]}"
        )

        if attempt < MAX_DETAIL_RETRY:
            try:
                await page.reload(wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            await asyncio.sleep(random.uniform(2.5, 4.0))
            continue   # ulang loop, ekstrak lagi setelah reload

        # Semua retry habis — putuskan berdasarkan best_extracted
        if best_extracted == 0:
            # Benar-benar kosong → skip screenshot
            print(f"      ⏭️  0 field berhasil — skip screenshot, status=no_data")
            logger.warning(
                f"NO_DATA | {product.get('link','')[:80]} | "
                f"keyword={product.get('keyword','')}"
            )
            product["scraped_at"]   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            product["status"]       = "no_data"
            product["screenshot"]   = ""
            product["harga_tinggi"] = "TIDAK"
            return product
        else:
            # Ada data parsial → tetap screenshot, tandai partial
            print(
                f"      🟡 Data parsial ({best_extracted}/{TOTAL_FIELDS}) "
                f"— screenshot dengan status=partial"
            )
            product["status"] = "partial"
            break   # keluar loop, lanjut ke screenshot di bawah

    # ── 6. Ambil screenshot (hanya dicapai jika ada data) ────────
    mahal     = is_expensive(product)
    ss_folder = SCREENSHOT_MAHAL if mahal else SCREENSHOT_FOLDER
    ss_path   = await take_screenshot(page, product, ss_folder)

    product["screenshot"]   = ss_path or ""
    product["harga_tinggi"] = "YA" if mahal else "TIDAK"
    product["scraped_at"]   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if mahal:
        print(f"      💰 Harga TINGGI ({product.get('price','?')}) → folder mahal/")
        notify_expensive(product)
        logger.warning(
            f"HARGA TINGGI | {product.get('price','?')} | "
            f"{product.get('title','')[:60]}"
        )

    return product


# ══════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════
COLS = [
    "No", "Keyword", "Judul Produk", "Harga", "Harga Tinggi?",
    "Rating", "Ulasan", "Terjual", "Stok", "Toko",
    "Status", "Link", "Screenshot", "Waktu Scan"
]
COL_WIDTHS = {
    "No": 5, "Keyword": 16, "Judul Produk": 42, "Harga": 18,
    "Harga Tinggi?": 14, "Rating": 9, "Ulasan": 11, "Terjual": 11,
    "Stok": 9, "Toko": 22, "Status": 12,
    "Link": 48, "Screenshot": 55, "Waktu Scan": 20,
}

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Warna status
STATUS_COLORS = {
    "ok"       : None,          # ikut warna baris normal
    "partial"  : "FFF9C4",      # kuning muda — data ada tapi tidak lengkap
    "no_data"  : "D3D3D3",      # abu-abu — 0 field berhasil
    "failed"   : "D3D3D3",      # abu-abu
    "blocked"  : "FFD700",      # kuning tua
    "no_link"  : "E0E0E0",      # abu-abu muda
}


def _write_header(ws, hdr_color="1F6AA5"):
    for col_idx, col_name in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color=hdr_color)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            COL_WIDTHS.get(col_name, 15)
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def _write_row(ws, row_idx: int, p: dict):
    mahal  = p.get("harga_tinggi") == "YA"
    status = p.get("status", "ok")

    # Tentukan warna background baris
    if mahal:
        bg = "FFB3B3"
    elif status in STATUS_COLORS and STATUS_COLORS[status]:
        bg = STATUS_COLORS[status]
    elif row_idx % 2 == 0:
        bg = "EBF3FB"
    else:
        bg = "FFFFFF"

    row_data = [
        row_idx,
        p.get("keyword", ""),
        p.get("title", "N/A"),
        p.get("price", "N/A"),
        p.get("harga_tinggi", "TIDAK"),
        p.get("rating", "N/A"),
        p.get("review_count", "N/A"),
        p.get("sold", "N/A"),
        p.get("stock", "N/A"),
        p.get("shop", "N/A"),
        status,
        p.get("link", ""),
        p.get("screenshot", ""),
        p.get("scraped_at", ""),
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell        = ws.cell(row=row_idx + 1, column=col_idx, value=value)
        cell.border = BORDER

        if mahal:
            cell.font = Font(
                name="Arial", size=10,
                bold=(col_idx in [3, 4, 5]),
                color="8B0000"
            )
            cell.fill = PatternFill("solid", start_color="FFB3B3")
        elif status in ("no_data", "failed", "no_link"):
            cell.font = Font(name="Arial", size=10, color="888888", italic=True)
            cell.fill = PatternFill("solid", start_color=bg)
        elif status == "partial":
            cell.font = Font(name="Arial", size=10, color="7D6608")
            cell.fill = PatternFill("solid", start_color=bg)
        elif status == "blocked":
            cell.font = Font(name="Arial", size=10, color="7D6000")
            cell.fill = PatternFill("solid", start_color=bg)
        else:
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)

        cell.alignment = Alignment(
            vertical="center", wrap_text=(col_idx in [3, 13])
        )

        # Kolom Harga Tinggi? → center
        if col_idx == 5:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Kolom Status → center + font bold
        if col_idx == 11:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if status == "ok":
                cell.font = Font(name="Arial", size=10, color="1E7E34", bold=True)
            elif status == "partial":
                cell.font = Font(name="Arial", size=10, color="7D6608", bold=True)
            elif status in ("no_data", "failed"):
                cell.font = Font(name="Arial", size=10, color="888888", bold=True, italic=True)
            elif status == "blocked":
                cell.font = Font(name="Arial", size=10, color="7D6000", bold=True)

        # Kolom Link → hyperlink style
        if col_idx == 12 and value:
            cell.font = Font(
                name="Arial", size=10,
                color="C0392B" if mahal else "1F6AA5",
                underline="single"
            )

        # Kolom Screenshot → link kecil
        if col_idx == 13 and value:
            cell.font = Font(
                name="Arial", size=9, color="555555", underline="single"
            )

        # Kolom Rating → angka desimal
        if col_idx == 6:
            try:
                cell.value         = float(value)
                cell.number_format = "0.0"
            except Exception:
                cell.value = value

    ws.row_dimensions[row_idx + 1].height = 25 if mahal else 22


def save_report(products: list[dict], keyword: str) -> str:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    safe_kw = sanitize_filename(keyword.lower())
    path    = _unique_path(
        os.path.join(OUTPUT_FOLDER, f"laporan_{safe_kw}.xlsx")
    )

    mahal_list    = [p for p in products if p.get("harga_tinggi") == "YA"]
    normal_list   = [p for p in products if p.get("harga_tinggi") != "YA"
                     and p.get("status", "ok") == "ok"]
    partial_list  = [p for p in products if p.get("status") == "partial"
                     and p.get("harga_tinggi") != "YA"]
    problem_list  = [p for p in products if p.get("status", "ok") in ("no_data", "failed", "no_link")
                     and p.get("harga_tinggi") != "YA"]

    print(f"   📊 {len(normal_list)} ok | "
          f"{len(mahal_list)} harga tinggi (> Rp{HARGA_THRESHOLD:,}) | "
          f"{len(partial_list)} parsial | "
          f"{len(problem_list)} no_data/gagal")

    wb = Workbook()
    ws = wb.active
    ws.title = keyword[:31]
    _write_header(ws)

    # Urutan: harga tinggi → normal → parsial → bermasalah
    sorted_products = mahal_list + normal_list + partial_list + problem_list
    for i, p in enumerate(sorted_products, 1):
        _write_row(ws, i, p)

    # Sheet ringkasan singkat
    ws2      = wb.create_sheet("Ringkasan")
    headers2 = [
        "Keyword", "Total", f"Harga > Rp{HARGA_THRESHOLD:,}",
        "Normal (ok)", "Parsial", "No Data/Gagal", "Tanggal"
    ]
    for col_idx, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col_idx, value=h).font = Font(bold=True)
    ws2["A2"] = keyword
    ws2["B2"] = len(products)
    ws2["C2"] = len(mahal_list)
    ws2["D2"] = len(normal_list)
    ws2["E2"] = len(partial_list)
    ws2["F2"] = len(problem_list)
    ws2["G2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if mahal_list:
        ws2["C2"].fill = PatternFill("solid", start_color="FFB3B3")
        ws2["C2"].font = Font(bold=True, color="8B0000")
    if partial_list:
        ws2["E2"].fill = PatternFill("solid", start_color="FFF9C4")
        ws2["E2"].font = Font(bold=True, color="7D6608")
    if problem_list:
        ws2["F2"].fill = PatternFill("solid", start_color="D3D3D3")
        ws2["F2"].font = Font(bold=True, color="888888")

    wb.save(path)
    print(f"   💾 Laporan → {path}")
    return path


# ══════════════════════════════════════════
#  RINGKASAN SEMUA KEYWORD
# ══════════════════════════════════════════
COLS_SUMMARY = [
    "Keyword", "Total Produk", "Harga Tinggi",
    "Normal (ok)", "Parsial", "No Data/Gagal", "% Harga Tinggi",
    "Harga Tertinggi", "Produk Termahal",
    "Link Produk Termahal", "Screenshot Termahal", "Waktu Scan"
]
COL_WIDTHS_SUMMARY = {
    "Keyword": 20, "Total Produk": 13, "Harga Tinggi": 14,
    "Normal (ok)": 12, "Parsial": 10, "No Data/Gagal": 14, "% Harga Tinggi": 16,
    "Harga Tertinggi": 18, "Produk Termahal": 45,
    "Link Produk Termahal": 50, "Screenshot Termahal": 50,
    "Waktu Scan": 20,
}

def save_summary(all_keyword_data: list[dict]):
    if not all_keyword_data:
        return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"

    for col_idx, col_name in enumerate(COLS_SUMMARY, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="2C3E50")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            COL_WIDTHS_SUMMARY.get(col_name, 15)
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    for row_idx, kw_data in enumerate(all_keyword_data, 1):
        products      = kw_data["products"]
        keyword       = kw_data["keyword"]
        total         = len(products)
        mahal_list    = [p for p in products if p.get("harga_tinggi") == "YA"]
        partial_list  = [p for p in products if p.get("status") == "partial"
                         and p.get("harga_tinggi") != "YA"]
        problem_list  = [p for p in products if p.get("status", "ok") in ("no_data", "failed", "no_link")
                         and p.get("harga_tinggi") != "YA"]
        normal_count  = total - len(mahal_list) - len(partial_list) - len(problem_list)
        pct           = round(len(mahal_list) / total * 100, 1) if total > 0 else 0

        top_p = (
            max(
                mahal_list,
                key=lambda p: extract_price_number(p.get("price", "0"))
            ) if mahal_list else {}
        )

        bg = (
            "FFB3B3" if pct >= 30
            else ("FFE5A0" if pct >= 10 else "B3FFB3")
        )

        row_data = [
            keyword, total, len(mahal_list),
            normal_count, len(partial_list), len(problem_list), f"{pct}%",
            top_p.get("price", "-"),
            top_p.get("title", "-"),
            top_p.get("link", "-"),
            top_p.get("screenshot", "-"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(
                vertical="center", wrap_text=(col_idx == 9)
            )
            cell.border = BORDER
            if col_idx in [2, 3, 4, 5, 6]:
                cell.font      = Font(name="Arial", size=10, bold=True)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if col_idx == 10 and value and value != "-":
                cell.font = Font(
                    name="Arial", size=10, color="1F6AA5", underline="single"
                )
        ws.row_dimensions[row_idx + 1].height = 22

    wb.save(SUMMARY_FILE)
    print(f"\n   📊 Ringkasan tersimpan: {SUMMARY_FILE}")


# ══════════════════════════════════════════
#  MAIN SCRAPER
# ══════════════════════════════════════════
async def run_scraper():
    keywords = load_keywords(KEYWORDS_FILE)
    if not keywords:
        return

    # [BARU] Setup signal handler graceful shutdown
    _setup_signal_handler()

    # [BARU] Inisialisasi ProxyManager dan AdaptiveRateLimit
    proxy_mgr = ProxyManager(PROXY_FILE)
    rate_rl   = AdaptiveRateLimit(base_min=1.0, base_max=2.5)

    mode = (
        f"maks {MAX_LOAD_MORE}x klik" if MAX_LOAD_MORE > 0
        else "klik semua tombol + fallback pagination"
    )
    print(f"📋 {len(keywords)} keyword dimuat.")
    print(f"⚙️  Mode         : {mode}")
    print(f"   💰 Threshold  : Rp{HARGA_THRESHOLD:,}")
    print(f"   ✅ Min field  : {MIN_FIELDS_OK}/{TOTAL_FIELDS} (langsung screenshot)")
    print(f"   ⏳ Data timeout: {WAIT_DATA_TIMEOUT}s")
    print(f"   🔄 Max retry  : {MAX_DETAIL_RETRY}x (jika field < {MIN_FIELDS_OK})")
    print(f"   ♾️  Spinner    : {SPINNER_TIMEOUT}s")
    print(f"   🌐 Proxy      : {'aktif' if USE_PROXY and proxy_mgr.proxies else 'direct'}")
    print(f"   📊 Backup JSONL: {'aktif' if USE_JSON_BACKUP else 'off'}")
    print(f"   📷 Screenshot : {'JPEG' if SCREENSHOT_JPEG else 'PNG'}")
    print(f"   📋 Log        : {LOG_FILE}")
    print()

    logger.info(
        f"MULAI — {len(keywords)} keyword | threshold Rp{HARGA_THRESHOLD:,}"
    )

    os.makedirs(SCREENSHOT_FOLDER, exist_ok=True)
    os.makedirs(SCREENSHOT_MAHAL, exist_ok=True)

    checkpoint    = checkpoint_load()
    done_keywords = set(checkpoint.get("done_keywords", []))
    if done_keywords:
        print(f"♻️  Resume — sudah selesai: {list(done_keywords)}\n")

    async with async_playwright() as p:
        browser = None
        is_cdp  = False
        try:
            print("   🔄 Mencoba koneksi ke Chrome asli (Anti-Bot Level Maksimal)...")
            browser = await p.chromium.connect_over_cdp("http://localhost:9222")
            is_cdp  = True
            print("   ✅ Berhasil terhubung ke Chrome asli!")
        except Exception:
            print("   ⚠️  Chrome asli tidak terdeteksi di port 9222.")
            print("   ⚠️  Fallback ke browser bawaan Playwright.")
            browser = await p.chromium.launch(
                headless=False,
                args=[
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-infobars",
                    "--disable-dev-shm-usage",
                    "--disable-http2",
                    "--ignore-certificate-errors",
                ],
            )

        if len(browser.contexts) > 0:
            context = browser.contexts[0]
            page    = await context.new_page()
            page.set_default_timeout(PAGE_TIMEOUT * 1000)
        else:
            # [MODIF] Pass proxy ke create_context
            context, page = await create_context(browser, proxy=proxy_mgr.next_proxy())

        context_ref      = [context]
        product_count    = 0
        all_saved        = []
        all_keyword_data = []

        for keyword in keywords:
            if keyword in done_keywords:
                print(f"⏭️  Skip '{keyword}' (sudah selesai)")
                continue

            print(f"\n{'═'*62}")
            print(f"🔍 Keyword: '{keyword}'")
            print(f"{'═'*62}")

            products, context, page = await scrape_all_pages(
                page, keyword, browser=browser, context_ref=context_ref, is_cdp=is_cdp
            )

            if not products:
                print(f"   ⚠️  Tidak ada produk, restart sesi...")
                if is_cdp:
                    try:
                        await page.close()
                    except Exception:
                        pass
                    await asyncio.sleep(random.uniform(5, 10))
                    # [FIX] Clear cookies agar session bersih
                    try:
                        await context_ref[0].clear_cookies()
                    except Exception:
                        pass
                    try:
                        page = await context_ref[0].new_page()
                    except Exception:
                        print("   ⚠️  Context mati, coba reconnect ke CDP...")
                        try:
                            browser = await p.chromium.connect_over_cdp("http://localhost:9222")
                            context_ref[0] = browser.contexts[0]
                            page = await context_ref[0].new_page()
                            page.set_default_timeout(PAGE_TIMEOUT * 1000)
                        except Exception as ex:
                            print(f"   ❌ Reconnect gagal: {ex} — skip keyword ini")
                            continue
                else:
                    try:
                        await context_ref[0].close()
                    except Exception:
                        pass
                    await asyncio.sleep(random.uniform(8, 15))
                    try:
                        context, page  = await create_context(browser, proxy=proxy_mgr.next_proxy())
                        context_ref[0] = context
                    except Exception as ex:
                        print(f"   ❌ Gagal buat context baru: {ex} — skip")
                        continue
                product_count = 0
                continue

            detailed_products = []
            total = len(products)
            
            sem = asyncio.Semaphore(MAX_CONCURRENT_TABS)
            next_restart = RESTART_EVERY

            tracker = ProgressTracker(total)  # [BARU]

            async def worker(product, p_idx):
                global _shutdown_requested
                if _shutdown_requested:
                    return product
                # Jitter awal di LUAR semaphore
                await rate_rl.wait()  # [BARU] adaptive delay
                await asyncio.sleep(random.uniform(0.3, 1.5))
                async with sem:
                    print(f"\n   [{p_idx}/{total}] {product['title'][:52]}...")
                    try:
                        detail_page = await context_ref[0].new_page()
                    except Exception:
                        return product
                    detail_page.set_default_timeout(PAGE_TIMEOUT * 1000)
                    try:
                        detailed = await scrape_product_detail(detail_page, product)
                        # [BARU] Adaptive rate limit feedback
                        st = detailed.get("status", "ok")
                        if st in ("ok", "partial"):
                            rate_rl.on_success()
                        else:
                            rate_rl.on_failure()
                        # [BARU] JSON backup incremental
                        backup_append(product.get("keyword", ""), detailed)
                        # [BARU] Progress dashboard
                        tracker.update(st)
                    except Exception as e:
                        print(f"      [Detail error: {e}]")
                        detailed = product
                        rate_rl.on_failure()
                    finally:
                        try:
                            await detail_page.close()
                        except Exception:
                            pass
                await human_delay(1.0, 2.0)
                return detailed

            for i in range(0, total, MAX_CONCURRENT_TABS):
                batch = products[i:i+MAX_CONCURRENT_TABS]
                
                if product_count >= next_restart:
                    next_restart += RESTART_EVERY
                    print(f"\n   🔄 Restart sesi (anti-throttle)...")
                    cooldown = random.uniform(COOLDOWN_ON_HANG, COOLDOWN_ON_HANG + 10)

                    if is_cdp:
                        try:
                            await page.close()
                        except Exception:
                            pass
                        # [FIX] Escalating cooldown — makin banyak produk makin lama istirahat
                        escalation   = 1.0 + (product_count / 80) * 0.6
                        cooldown_eff = cooldown * min(escalation, 4.0)
                        print(f"   💤 Cooldown {cooldown_eff:.0f}s (x{min(escalation,4.0):.1f} setelah {product_count} produk)...")
                        # [FIX] Clear cookies — hapus session flag bot Tokopedia
                        try:
                            await context_ref[0].clear_cookies()
                            print("   🧹 Cookies di-clear")
                        except Exception:
                            pass
                        await asyncio.sleep(cooldown_eff)
                        try:
                            page = await context_ref[0].new_page()
                        except Exception:
                            print("   ⚠️  Context mati — reconnect ke CDP...")
                            try:
                                browser = await p.chromium.connect_over_cdp("http://localhost:9222")
                                context_ref[0] = browser.contexts[0]
                                page = await context_ref[0].new_page()
                                page.set_default_timeout(PAGE_TIMEOUT * 1000)
                            except Exception as ex:
                                print(f"   ❌ Reconnect gagal: {ex}")
                    else:
                        try:
                            await context_ref[0].close()
                        except Exception:
                            pass
                        escalation   = 1.0 + (product_count / 80) * 0.6
                        cooldown_eff = cooldown * min(escalation, 4.0)
                        print(f"   💤 Cooldown {cooldown_eff:.0f}s (x{min(escalation,4.0):.1f})...")
                        await asyncio.sleep(cooldown_eff)
                        context, page = await create_context(browser, proxy=proxy_mgr.next_proxy())
                        context_ref[0] = context

                    print(f"   ✅ Sesi baru siap")

                tasks = [worker(p, i + j + 1) for j, p in enumerate(batch)]

                # [FIX] return_exceptions=True agar 1 task gagal tidak cancel semua
                batch_results = await asyncio.gather(*tasks, return_exceptions=True)
                # Ganti exception dengan produk original (tidak hilang dari laporan)
                batch_results = [
                    products[i + j] if isinstance(r, Exception) else r
                    for j, r in enumerate(batch_results)
                ]

                detailed_products.extend(batch_results)
                product_count += len(batch)

                # [BARU] Cek shutdown setelah tiap batch
                if _shutdown_requested:
                    print("   🛑 Shutdown diminta — menyimpan progress...")
                    break

            path = save_report(detailed_products, keyword)
            all_saved.append((keyword, path))
            all_keyword_data.append({
                "keyword" : keyword,
                "products": detailed_products
            })

            done_keywords.add(keyword)
            checkpoint_save({"done_keywords": list(done_keywords)})
            print(
                f"   💾 Checkpoint: {len(done_keywords)}/{len(keywords)} keyword"
            )

            if _shutdown_requested:
                break
            await human_delay(5.0, 10.0)

        # [BARU] Retry Queue — produk blocked/no_data dicoba ulang
        if USE_RETRY_QUEUE and not _shutdown_requested and all_keyword_data:
            all_failed = [
                (kw_d["keyword"], p)
                for kw_d in all_keyword_data
                for p in kw_d["products"]
                if p.get("status") in ("blocked", "no_data", "failed")
            ]
            if all_failed:
                print(f"\n{'\u2550'*62}")
                print(f"🔄 Retry Queue: {len(all_failed)} produk gagal — restart sesi dulu...")
                cooldown = random.uniform(20, 35)
                print(f"   💤 Cooldown {cooldown:.0f}s sebelum retry...")
                await asyncio.sleep(cooldown)
                if not is_cdp:
                    try:
                        await context_ref[0].close()
                    except Exception:
                        pass
                    ctx_r, page = await create_context(browser, proxy=proxy_mgr.next_proxy())
                    context_ref[0] = ctx_r
                for kw_label, p in all_failed:
                    if _shutdown_requested:
                        break
                    await asyncio.sleep(random.uniform(5, 10))
                    try:
                        rp = await context_ref[0].new_page()
                        rp.set_default_timeout(PAGE_TIMEOUT * 1000)
                        retried = await scrape_product_detail(rp, p)
                        await rp.close()
                        # Update di all_keyword_data
                        for kd in all_keyword_data:
                            if kd["keyword"] == kw_label:
                                for idx, prod in enumerate(kd["products"]):
                                    if prod.get("link") == retried.get("link"):
                                        kd["products"][idx] = retried
                                        backup_append(kw_label, retried)
                                        break
                        st = retried.get("status", "?")
                        print(f"   🔄 Retry '{p.get('title','')[:40]}' → {st}")
                    except Exception as e:
                        print(f"   ⚠️  Retry error: {e}")
                # Simpan ulang laporan setelah retry
                for kd in all_keyword_data:
                    save_report(kd["products"], kd["keyword"])
                print(f"💾 Laporan di-update setelah retry.")

        try:
            await context_ref[0].close()
        except Exception:
            pass
        try:
            await browser.close()
        except Exception:
            pass

    checkpoint_clear()
    save_summary(all_keyword_data)

    total_semua   = sum(len(d["products"]) for d in all_keyword_data)
    total_mahal   = sum(
        len([p for p in d["products"] if p.get("harga_tinggi") == "YA"])
        for d in all_keyword_data
    )
    total_partial = sum(
        len([p for p in d["products"] if p.get("status") == "partial"])
        for d in all_keyword_data
    )
    total_nodata  = sum(
        len([p for p in d["products"] if p.get("status", "ok") in ("no_data", "failed", "no_link")])
        for d in all_keyword_data
    )

    print(f"\n{'═'*62}")
    print(f"🎉 Selesai!")
    print(f"   📦 Total produk     : {total_semua}")
    print(f"   💰 Harga tinggi     : {total_mahal} (> Rp{HARGA_THRESHOLD:,})")
    print(f"   ✅ Normal (ok)      : {total_semua - total_mahal - total_partial - total_nodata}")
    print(f"   🟡 Parsial          : {total_partial} (< {MIN_FIELDS_OK}/{TOTAL_FIELDS} field)")
    print(f"   ⚠️  No data/gagal   : {total_nodata}")
    print(f"   📊 Ringkasan        : {SUMMARY_FILE}")
    print(f"   📋 Log              : {LOG_FILE}")
    print(f"   📷 Screenshot semua : {SCREENSHOT_FOLDER}/")
    print(f"   📷 Screenshot mahal : {SCREENSHOT_MAHAL}/")
    print(f"\n   Laporan per keyword:")
    for kw, path in all_saved:
        print(f"      📁 '{kw}' → {path}")
    print(f"{'═'*62}")

    logger.info(
        f"SELESAI — total={total_semua} | mahal={total_mahal} | "
        f"normal={total_semua - total_mahal - total_partial - total_nodata} | "
        f"partial={total_partial} | nodata={total_nodata}"
    )


if __name__ == "__main__":
    asyncio.run(run_scraper())