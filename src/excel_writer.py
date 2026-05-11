import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import HARGA_THRESHOLD, SUMMARY_FILE, OUTPUT_FOLDER
from utils import extract_price_number, sanitize_filename, _unique_path

# ══════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════
COLS = [
    "No", "Keyword", "Judul Produk", "Harga", "Harga Tinggi?",
    "Rating", "Ulasan", "Terjual", "Stok", "Toko",
    "Status", "Link", "Screenshot", "Waktu Scan"
]
COL_WIDTHS = {
    "No": 5, "Keyword": 16, "Judul Produk": 42, "Harga": 18,
    "Harga Tinggi?": 14, "Rating": 9, "Ulasan": 11, "Terjual": 11,
    "Stok": 9, "Toko": 22, "Status": 12,
    "Link": 48, "Screenshot": 55, "Waktu Scan": 20,
}

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Warna status
STATUS_COLORS = {
    "ok"       : None,          # ikut warna baris normal
    "partial"  : "FFF9C4",      # kuning muda — data ada tapi tidak lengkap
    "no_data"  : "D3D3D3",      # abu-abu — 0 field berhasil
    "failed"   : "D3D3D3",      # abu-abu
    "blocked"  : "FFD700",      # kuning tua
    "no_link"  : "E0E0E0",      # abu-abu muda
}

def _write_header(ws, hdr_color="1F6AA5"):
    for col_idx, col_name in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color=hdr_color)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            COL_WIDTHS.get(col_name, 15)
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

def _write_row(ws, row_idx: int, p: dict):
    mahal  = p.get("harga_tinggi") == "YA"
    status = p.get("status", "ok")

    # Tentukan warna background baris
    if mahal:
        bg = "FFB3B3"
    elif status in STATUS_COLORS and STATUS_COLORS[status]:
        bg = STATUS_COLORS[status]
    elif row_idx % 2 == 0:
        bg = "EBF3FB"
    else:
        bg = "FFFFFF"

    row_data = [
        row_idx,
        p.get("keyword", ""),
        p.get("title", "N/A"),
        p.get("price", "N/A"),
        p.get("harga_tinggi", "TIDAK"),
        p.get("rating", "N/A"),
        p.get("review_count", "N/A"),
        p.get("sold", "N/A"),
        p.get("stock", "N/A"),
        p.get("shop", "N/A"),
        status,
        p.get("link", ""),
        p.get("screenshot", ""),
        p.get("scraped_at", ""),
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell        = ws.cell(row=row_idx + 1, column=col_idx, value=value)
        cell.border = BORDER

        if mahal:
            cell.font = Font(
                name="Arial", size=10,
                bold=(col_idx in [3, 4, 5]),
                color="8B0000"
            )
            cell.fill = PatternFill("solid", start_color="FFB3B3")
        elif status in ("no_data", "failed", "no_link"):
            cell.font = Font(name="Arial", size=10, color="888888", italic=True)
            cell.fill = PatternFill("solid", start_color=bg)
        elif status == "partial":
            cell.font = Font(name="Arial", size=10, color="7D6608")
            cell.fill = PatternFill("solid", start_color=bg)
        elif status == "blocked":
            cell.font = Font(name="Arial", size=10, color="7D6000")
            cell.fill = PatternFill("solid", start_color=bg)
        else:
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)

        cell.alignment = Alignment(
            vertical="center", wrap_text=(col_idx in [3, 13])
        )

        if col_idx == 5:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if col_idx == 11:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if status == "ok":
                cell.font = Font(name="Arial", size=10, color="1E7E34", bold=True)
            elif status == "partial":
                cell.font = Font(name="Arial", size=10, color="7D6608", bold=True)
            elif status in ("no_data", "failed"):
                cell.font = Font(name="Arial", size=10, color="888888", bold=True, italic=True)
            elif status == "blocked":
                cell.font = Font(name="Arial", size=10, color="7D6000", bold=True)

        if col_idx == 12 and value:
            cell.font = Font(
                name="Arial", size=10,
                color="C0392B" if mahal else "1F6AA5",
                underline="single"
            )

        if col_idx == 13 and value:
            cell.font = Font(
                name="Arial", size=9, color="555555", underline="single"
            )

        if col_idx == 6:
            try:
                cell.value         = float(value)
                cell.number_format = "0.0"
            except Exception:
                cell.value = value

    ws.row_dimensions[row_idx + 1].height = 25 if mahal else 22


def save_report(products: list[dict], keyword: str) -> str:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    safe_kw = sanitize_filename(keyword.lower())
    path    = _unique_path(
        os.path.join(OUTPUT_FOLDER, f"laporan_{safe_kw}.xlsx")
    )

    mahal_list    = [p for p in products if p.get("harga_tinggi") == "YA"]
    normal_list   = [p for p in products if p.get("harga_tinggi") != "YA"
                     and p.get("status", "ok") == "ok"]
    partial_list  = [p for p in products if p.get("status") == "partial"
                     and p.get("harga_tinggi") != "YA"]
    problem_list  = [p for p in products if p.get("status", "ok") in ("no_data", "failed", "no_link")
                     and p.get("harga_tinggi") != "YA"]

    print(f"   📊 {len(normal_list)} ok | "
          f"{len(mahal_list)} harga tinggi (> Rp{HARGA_THRESHOLD:,}) | "
          f"{len(partial_list)} parsial | "
          f"{len(problem_list)} no_data/gagal")

    wb = Workbook()
    ws = wb.active
    ws.title = keyword[:31]
    _write_header(ws)

    sorted_products = mahal_list + normal_list + partial_list + problem_list
    for i, p in enumerate(sorted_products, 1):
        _write_row(ws, i, p)

    ws2      = wb.create_sheet("Ringkasan")
    headers2 = [
        "Keyword", "Total", f"Harga > Rp{HARGA_THRESHOLD:,}",
        "Normal (ok)", "Parsial", "No Data/Gagal", "Tanggal"
    ]
    for col_idx, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col_idx, value=h).font = Font(bold=True)
    ws2["A2"] = keyword
    ws2["B2"] = len(products)
    ws2["C2"] = len(mahal_list)
    ws2["D2"] = len(normal_list)
    ws2["E2"] = len(partial_list)
    ws2["F2"] = len(problem_list)
    ws2["G2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if mahal_list:
        ws2["C2"].fill = PatternFill("solid", start_color="FFB3B3")
        ws2["C2"].font = Font(bold=True, color="8B0000")
    if partial_list:
        ws2["E2"].fill = PatternFill("solid", start_color="FFF9C4")
        ws2["E2"].font = Font(bold=True, color="7D6608")
    if problem_list:
        ws2["F2"].fill = PatternFill("solid", start_color="D3D3D3")
        ws2["F2"].font = Font(bold=True, color="888888")

    wb.save(path)
    print(f"   💾 Laporan → {path}")
    return path


COLS_SUMMARY = [
    "Keyword", "Total Produk", "Harga Tinggi",
    "Normal (ok)", "Parsial", "No Data/Gagal", "% Harga Tinggi",
    "Harga Tertinggi", "Produk Termahal",
    "Link Produk Termahal", "Screenshot Termahal", "Waktu Scan"
]
COL_WIDTHS_SUMMARY = {
    "Keyword": 20, "Total Produk": 13, "Harga Tinggi": 14,
    "Normal (ok)": 12, "Parsial": 10, "No Data/Gagal": 14, "% Harga Tinggi": 16,
    "Harga Tertinggi": 18, "Produk Termahal": 45,
    "Link Produk Termahal": 50, "Screenshot Termahal": 50,
    "Waktu Scan": 20,
}

def save_summary(all_keyword_data: list[dict]):
    if not all_keyword_data:
        return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"

    for col_idx, col_name in enumerate(COLS_SUMMARY, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="2C3E50")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            COL_WIDTHS_SUMMARY.get(col_name, 15)
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    for row_idx, kw_data in enumerate(all_keyword_data, 1):
        products      = kw_data["products"]
        keyword       = kw_data["keyword"]
        total         = len(products)
        mahal_list    = [p for p in products if p.get("harga_tinggi") == "YA"]
        partial_list  = [p for p in products if p.get("status") == "partial"
                         and p.get("harga_tinggi") != "YA"]
        problem_list  = [p for p in products if p.get("status", "ok") in ("no_data", "failed", "no_link")
                         and p.get("harga_tinggi") != "YA"]
        normal_count  = total - len(mahal_list) - len(partial_list) - len(problem_list)
        pct           = round(len(mahal_list) / total * 100, 1) if total > 0 else 0

        top_p = (
            max(
                mahal_list,
                key=lambda p: extract_price_number(p.get("price", "0"))
            ) if mahal_list else {}
        )

        bg = (
            "FFB3B3" if pct >= 30
            else ("FFE5A0" if pct >= 10 else "B3FFB3")
        )

        row_data = [
            keyword, total, len(mahal_list),
            normal_count, len(partial_list), len(problem_list), f"{pct}%",
            top_p.get("price", "-"),
            top_p.get("title", "-"),
            top_p.get("link", "-"),
            top_p.get("screenshot", "-"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(
                vertical="center", wrap_text=(col_idx == 9)
            )
            cell.border = BORDER
            if col_idx in [2, 3, 4, 5, 6]:
                cell.font      = Font(name="Arial", size=10, bold=True)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if col_idx == 10 and value and value != "-":
                cell.font = Font(
                    name="Arial", size=10, color="1F6AA5", underline="single"
                )
        ws.row_dimensions[row_idx + 1].height = 22

    wb.save(SUMMARY_FILE)
    print(f"\n   📊 Ringkasan tersimpan: {SUMMARY_FILE}")
