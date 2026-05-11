"""
SiPantau Local Agent
====================
Jalankan file ini di komputer Anda.
Website SiPantau akan otomatis terhubung ke agent ini
dan memerintahkan scraping Tokopedia.

Cara menjalankan:
  - Double-click: jalankan_agent.bat
  - Atau manual:  python agent.py
"""

import asyncio
import os
import sys
import uuid
import threading
import time
import re
import requests
import subprocess
import random

from datetime import datetime
from typing import Optional, Dict, Any

# PENTING: Paksa Playwright menyimpan/mencari browser di folder permanen,
# BUKAN di folder temporary (_MEI) milik PyInstaller yang selalu berubah/terhapus.
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(os.path.expanduser("~"), "AppData", "Local", "ms-playwright")

# Pastikan working directory = folder tempat .exe berada (bukan folder temporary PyInstaller)
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))
os.chdir(application_path)

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

AGENT_VERSION    = "1.1.0"
AGENT_PORT       = 7777
SIPANTAU_WEB_URL = "http://localhost:3000"  # Ganti jika web di-host di Vercel/server lain

# ── Status browser (untuk web UI) ──
_browser_status  = "checking"   # "checking" | "downloading" | "ready" | "error"
_browser_message = "Memeriksa browser..."
_browser_ready   = threading.Event()

def ensure_browser_installed():
    """Install Chromium otomatis + buka browser ke SiPantau setelah siap."""
    global _browser_status, _browser_message
    import subprocess, webbrowser
    try:
        from playwright._impl._driver import compute_driver_executable
        driver_executable, driver_cli = compute_driver_executable()

        print("\n" + "═"*55)
        print("  SiPantau Agent — Mempersiapkan browser...")
        print("  (Hanya butuh beberapa menit di pertama kali)")
        print("═"*55)

        _browser_status  = "downloading"
        _browser_message = "Menyiapkan browser otomatis..."

        proc = subprocess.Popen(
            [driver_executable, driver_cli, "install", "chromium"],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace"
        )
        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue
            pct = re.search(r"(\d+)\s*%", line)
            if pct:
                msg = f"Mengunduh browser... {pct.group(1)}%"
                print(f"  → {msg}")
                _browser_message = msg
            elif any(k in line.lower() for k in ("downloaded", "installing", "unzip")):
                _browser_message = "Menginstall browser..."
                print(f"  → {line}")
        proc.wait()

        if proc.returncode == 0:
            _browser_status  = "ready"
            _browser_message = "Browser siap."
            _browser_ready.set()

            print()
            print("═"*55)
            print("  ✅ Browser siap!")
            print(f"  → Membuka halaman SiPantau...")
            print("═"*55)

            # Beri waktu server naik, lalu buka browser otomatis
            import time as _time
            _time.sleep(1.5)
            try:
                webbrowser.open(SIPANTAU_WEB_URL)
                print(f"  ✓ Halaman SiPantau terbuka di browser.")
                print(f"  Jika tidak terbuka otomatis, buka manual: {SIPANTAU_WEB_URL}")
            except Exception:
                print(f"  Buka manual di browser: {SIPANTAU_WEB_URL}")
            print()
            print("  Agent aktif. JANGAN tutup jendela hitam ini.")
            print("  Tekan Ctrl+C untuk menghentikan agent.")
            print("═"*55 + "\n")
        else:
            _browser_status  = "error"
            _browser_message = "Gagal menyiapkan browser. Coba jalankan ulang."
            _browser_ready.set()
            print("  ❌ Gagal. Coba tutup dan buka ulang SiPantau_Agent.exe.")

    except Exception as e:
        _browser_status  = "error"
        _browser_message = f"Error: {e}"
        _browser_ready.set()
        print(f"  Error: {e}")

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    threading.Thread(target=ensure_browser_installed, daemon=True).start()
    yield

app = FastAPI(title="SiPantau Local Agent", version=AGENT_VERSION, lifespan=lifespan)

# CORS — izinkan request dari semua origin (termasuk Vercel/Netlify)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

os.makedirs("output", exist_ok=True)
os.makedirs("output/checkpoint", exist_ok=True)

# ══════════════════════════════════════════
#  PENYIMPANAN STATUS JOB
# ══════════════════════════════════════════
jobs: Dict[str, Dict[str, Any]] = {}
# Format:
# {
#   "job_id": {
#     "status":      "queued" | "running" | "done" | "error" | "cancelled",
#     "keyword":     str,
#     "max_pages":   int,
#     "username":    str,
#     "backend_url": str,
#     "total":       int,       # produk ditemukan
#     "message":     str,       # pesan progress terbaru
#     "results":     list,      # produk hasil scraping
#     "file_excel":  str,       # nama file Excel
#     "started_at":  str,
#     "finished_at": str,
#   }
# }

# Semaphore — hanya 1 scraping bersamaan per agent
_scrape_lock = asyncio.Semaphore(1)
_event_loop: Optional[asyncio.AbstractEventLoop] = None

# ══════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════
class StartScrapeRequest(BaseModel):
    keyword:     str
    max_pages:   int    = 3
    target_product_count: int = 50  # jumlah produk BARU per sesi (max 50)
    harga_threshold: int = 350000
    username:    str    = "unknown"
    backend_url: str    = ""      # URL backend SiPantau untuk kirim hasil

class CancelRequest(BaseModel):
    job_id: str

# ══════════════════════════════════════════
#  ENDPOINTS
# ══════════════════════════════════════════
@app.get("/ping")
def ping():
    """Cek apakah agent aktif — dipanggil oleh web setiap kali halaman scraping dibuka."""
    active_jobs = [j for j in jobs.values() if j["status"] in ("queued", "running")]
    return {
        "status":          "ok",
        "version":         AGENT_VERSION,
        "active_jobs":     len(active_jobs),
        "agent":           "SiPantau Local Agent",
        "browser_ready":   (_browser_status == "ready"),
        "browser_status":  _browser_status,
        "browser_message": _browser_message,
    }

@app.post("/scrape")
async def start_scrape(req: StartScrapeRequest):
    """Mulai job scraping baru. Return job_id untuk polling status."""
    if not req.keyword.strip():
        raise HTTPException(status_code=400, detail="Keyword tidak boleh kosong")

    # Blokir jika browser belum siap
    if _browser_status not in ("ready", "error"):
        raise HTTPException(
            status_code=503,
            detail=f"Browser belum siap: {_browser_message}. Tunggu sebentar lalu coba lagi."
        )
    if _browser_status == "error":
        raise HTTPException(
            status_code=503,
            detail="Browser gagal diinstall. Coba tutup dan buka ulang SiPantau_Agent.exe."
        )

    # Cek apakah ada job yang sedang berjalan
    running = [j for j in jobs.values() if j["status"] == "running"]
    if running:
        raise HTTPException(
            status_code=429,
            detail="Agent sedang memproses scraping lain. Tunggu hingga selesai."
        )

    job_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]

    jobs[job_id] = {
        "status":      "queued",
        "keyword":     req.keyword.strip(),
        "max_pages":   req.max_pages,
        "target_product_count": req.target_product_count,
        "harga_threshold": req.harga_threshold,
        "username":    req.username,
        "backend_url": req.backend_url,
        "total":       0,
        "message":     "Antri — menunggu giliran...",
        "results":     [],
        "file_excel":  "",
        "started_at":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "finished_at": "",
    }

    # Jalankan di background (non-blocking)
    asyncio.create_task(_run_scrape_job(job_id))

    return {"job_id": job_id, "status": "queued", "message": "Job dimulai"}

@app.get("/status/{job_id}")
def get_status(job_id: str):
    """Polling status job scraping."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job tidak ditemukan")
    job = jobs[job_id]
    return {
        "job_id":      job_id,
        "status":      job["status"],
        "keyword":     job["keyword"],
        "total":       job["total"],
        "message":     job["message"],
        "file_excel":  job["file_excel"],
        "started_at":  job["started_at"],
        "finished_at": job["finished_at"],
    }

@app.get("/results/{job_id}")
def get_results(job_id: str):
    """Ambil hasil scraping lengkap."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job tidak ditemukan")
    job = jobs[job_id]
    if job["status"] not in ("done", "running", "error"):
        raise HTTPException(status_code=400, detail=f"Job belum selesai (status: {job['status']})")
    return {
        "job_id":        job_id,
        "keyword":       job["keyword"],
        "total":         job["total"],
        "results":       job["results"],
        "file_excel":    job["file_excel"],
        "upload_status": job.get("upload_status", ""),
    }

@app.get("/download/{job_id}")
def download_excel(job_id: str):
    """Download file Excel hasil scraping langsung dari Agent."""
    from fastapi.responses import FileResponse
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job tidak ditemukan")
    job = jobs[job_id]
    fname = job.get("file_excel", "")
    fpath = job.get("excel_path", "")  # path absolut dari src/output/
    if not fname or not fpath:
        raise HTTPException(status_code=404, detail="File Excel belum tersedia")
    if not os.path.exists(fpath):
        raise HTTPException(status_code=404, detail="File tidak ditemukan di disk")
    return FileResponse(
        path=fpath, filename=fname,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/open-output-folder")
def open_output_folder():
    """Buka folder output (Excel & Screenshot) di File Explorer Windows."""
    out_dir = os.path.abspath("output")
    if not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    try:
        os.startfile(out_dir)
        return {"success": True, "message": "Folder dibuka"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal membuka folder: {e}")

@app.post("/cancel/{job_id}")
def cancel_job(job_id: str):
    """Batalkan job yang sedang antri (tidak bisa batalkan yang sedang running)."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job tidak ditemukan")
    job = jobs[job_id]
    if job["status"] == "running":
        raise HTTPException(status_code=400, detail="Tidak bisa batalkan job yang sedang berjalan")
    if job["status"] == "queued":
        job["status"] = "cancelled"
        job["message"] = "Dibatalkan oleh user"
    return {"job_id": job_id, "status": job["status"]}

@app.get("/jobs")
def list_jobs():
    """Daftar semua job (max 20 terbaru)."""
    job_list = [
        {"job_id": jid, "status": j["status"], "keyword": j["keyword"],
         "total": j["total"], "started_at": j["started_at"]}
        for jid, j in list(jobs.items())[-20:]
    ]
    return {"jobs": list(reversed(job_list))}

# ══════════════════════════════════════════
#  SCRAPING ENGINE
# ══════════════════════════════════════════
async def _run_scrape_job(job_id: str):
    """
    Background task: jalankan scraping langsung (tanpa subprocess).
    Semua kode scraper sudah di-embed di dalam exe ini — tidak butuh src/ eksternal.
    Logika identik dengan src/main.py / run.bat.
    """
    job = jobs[job_id]
    job["status"]  = "running"
    job["message"] = "Memulai browser..."

    # Set env vars agar scraper.config membacanya dengan benar
    os.environ["SCRAPER_SCROLL"]    = str(job["target_product_count"])
    os.environ["SCRAPER_THRESHOLD"] = str(job["harga_threshold"])

    # Import modul scraper yang sudah di-embed
    from scraper.config       import PAGE_TIMEOUT, MAX_CONCURRENT_TABS, RESTART_EVERY, COOLDOWN_ON_HANG
    from scraper.browser_manager import AdaptiveRateLimit, create_context
    from scraper.scraper_core import scrape_all_pages, scrape_product_detail, _apply_stealth
    from scraper.excel_writer  import save_report
    from scraper.utils         import human_delay, ProgressTracker, backup_append, notify_expensive
    from scraper.proxy_manager import ProxyManager
    from playwright.async_api  import async_playwright

    proxy_mgr = ProxyManager()
    rate_rl   = AdaptiveRateLimit()

    # Output ke folder di samping exe (atau working dir)
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, "screenshots"), exist_ok=True)
    os.makedirs(os.path.join(output_dir, "screenshots", "mahal"), exist_ok=True)

    try:
        async with async_playwright() as p:
            browser = None
            is_cdp  = False
            try:
                job["message"] = "Mencoba koneksi ke Chrome asli (CDP)..."
                browser = await p.chromium.connect_over_cdp("http://localhost:9222")
                is_cdp  = True
                job["message"] = "✅ Chrome asli terhubung!"
                print("[Agent] Terhubung ke Chrome via CDP")
            except Exception:
                job["message"] = "Membuka Chrome asli secara otomatis..."
                print("[Agent] CDP tidak ditemukan. Mencoba membuka Chrome asli...")

                chrome_paths = [
                    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                    os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe")
                ]
                exe_path = next((path for path in chrome_paths if os.path.exists(path)), None)
                
                if exe_path:
                    temp_dir = os.path.join(os.environ.get("TEMP", "C:\\temp"), "chrome_scraping_profile")
                    # Gunakan CMD 'start' agar Chrome berjalan benar-benar terpisah (detached)
                    # Ini mencegah Chrome freeze/layar putih akibat terikat ke proses agent.
                    cmd = f'start "" "{exe_path}" --remote-debugging-port=9222 --user-data-dir="{temp_dir}"'
                    os.system(cmd)
                    await asyncio.sleep(4) # Tunggu Chrome terbuka sempurna
                    try:
                        browser = await p.chromium.connect_over_cdp("http://localhost:9222")
                        is_cdp = True
                        print("[Agent] Berhasil membuka dan terhubung ke Chrome asli via CDP!")
                    except Exception as e:
                        print(f"[Agent] Gagal konek ke Chrome asli yang baru dibuka: {e}")
                
                if not browser:
                    job["message"] = "Chrome tidak terdeteksi — fallback ke browser Playwright..."
                    print("[Agent] Fallback ke Playwright browser (Risiko terdeteksi bot lebih tinggi)")
                    browser = await p.chromium.launch(
                        headless=False,
                        args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                              "--disable-infobars", "--disable-dev-shm-usage"],
                    )

            if browser.contexts:
                context = browser.contexts[0]
                page    = await context.new_page()
                page.set_default_timeout(PAGE_TIMEOUT * 1000)
            else:
                context, page = await create_context(browser, proxy=await proxy_mgr.get_valid_proxy())

            context_ref   = [context]
            product_count = 0
            all_products  = []

            keyword = job["keyword"]
            print(f"\n{'═'*60}\n🔍 Keyword: '{keyword}'\n{'═'*60}")
            job["message"] = f"Membuka Tokopedia untuk '{keyword}'..."

            scrape_products, context, page = await scrape_all_pages(
                page, keyword,
                browser=browser,
                context_ref=context_ref,
                is_cdp=is_cdp,
                target_product_count=job["target_product_count"],
                harga_threshold=job.get("harga_threshold", 0),
            )

            if not scrape_products:
                job["message"] = "Tidak ada produk ditemukan."
                job["status"]  = "done"
                job["total"]   = 0
                job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                return

            total = len(scrape_products)
            job["total"]   = total
            job["message"] = f"{total} produk listing ditemukan. Membuka detail..."
            print(f"[Agent] {total} produk ditemukan. Mulai detail scraping...")

            detailed_products = []
            sem = asyncio.Semaphore(MAX_CONCURRENT_TABS)
            next_restart = RESTART_EVERY

            tracker = ProgressTracker(total)

            async def worker(product, p_idx):
                async with sem:
                    job["message"] = f"[{p_idx}/{total}] {product.get('title','')[:50]}..."
                    print(f"\n   [{p_idx}/{total}] {product.get('title','')[:52]}...")
                    try:
                        detail_page = await context_ref[0].new_page()
                        await detail_page.bring_to_front() # Fokuskan tab agar tidak kena throttle Chrome
                        detail_page.set_default_timeout(PAGE_TIMEOUT * 1000)
                        if not is_cdp:
                            await _apply_stealth(detail_page) # Jangan inject stealth ke Chrome asli
                    except Exception:
                        return product
                    try:
                        detailed = await scrape_product_detail(detail_page, product, job.get("harga_threshold", 0))
                        st = detailed.get("status", "ok")
                        if st in ("ok", "partial"):
                            rate_rl.on_success()
                        else:
                            rate_rl.on_failure()
                        backup_append(keyword, detailed)
                        tracker.update(st)
                    except Exception as e:
                        print(f"      [Detail error: {e}]")
                        detailed = product
                        rate_rl.on_failure()
                    finally:
                        try:
                            await detail_page.close()
                        except Exception:
                            pass
                await human_delay(1.0, 2.0)
                return detailed

            for i in range(0, total, MAX_CONCURRENT_TABS):
                batch = scrape_products[i:i+MAX_CONCURRENT_TABS]

                if product_count >= next_restart:
                    next_restart += RESTART_EVERY
                    print(f"\n   🔄 Restart sesi (anti-throttle)...")
                    cooldown = random.uniform(COOLDOWN_ON_HANG, COOLDOWN_ON_HANG + 10)
                    if is_cdp:
                        try:
                            await page.close()
                        except Exception:
                            pass
                        escalation   = 1.0 + (product_count / 80) * 0.6
                        cooldown_eff = cooldown * min(escalation, 4.0)
                        print(f"   💤 Cooldown {cooldown_eff:.0f}s...")
                        try:
                            await context_ref[0].clear_cookies()
                        except Exception:
                            pass
                        await asyncio.sleep(cooldown_eff)
                        try:
                            page = await context_ref[0].new_page()
                        except Exception:
                            try:
                                browser = await p.chromium.connect_over_cdp("http://localhost:9222")
                                context_ref[0] = browser.contexts[0]
                                page = await context_ref[0].new_page()
                                page.set_default_timeout(PAGE_TIMEOUT * 1000)
                            except Exception as ex:
                                print(f"   ❌ Reconnect gagal: {ex}")
                    else:
                        try:
                            await context_ref[0].close()
                        except Exception:
                            pass
                        escalation   = 1.0 + (product_count / 80) * 0.6
                        cooldown_eff = cooldown * min(escalation, 4.0)
                        await asyncio.sleep(cooldown_eff)
                        context, page = await create_context(browser, proxy=await proxy_mgr.get_valid_proxy())
                        context_ref[0] = context

                tasks = []
                for j, prod in enumerate(batch):
                    await rate_rl.wait()
                    await asyncio.sleep(random.uniform(0.3, 1.5))
                    tasks.append(worker(prod, i + j + 1))

                batch_results = await asyncio.gather(*tasks, return_exceptions=True)
                batch_results = [
                    scrape_products[i + j] if isinstance(r, Exception) else r
                    for j, r in enumerate(batch_results)
                ]
                detailed_products.extend(batch_results)
                product_count += len(batch)

            try:
                await context_ref[0].close()
            except Exception:
                pass
            try:
                await browser.close()
            except Exception:
                pass

        # Simpan ke Excel (menggunakan excel_writer dari scraper asli)
        # Gabungkan produk sesi ini dengan semua sesi sebelumnya (mode kumulatif)
        job["message"] = "Menyimpan Excel (kumulatif)..."
        from scraper.utils import backup_load_all_products
        all_previous = backup_load_all_products(keyword)
        # Hindari duplikat: produk baru sudah di-append ke backup oleh worker
        # Produk lama di backup = semua sesi, produk baru sudah termasuk
        # Kita pakai all_previous yang sudah berisi SEMUA produk termasuk sesi ini
        if all_previous:
            products_for_excel = all_previous
            print(f"[Agent] Membuat Excel kumulatif: {len(products_for_excel)} produk total ({len(detailed_products)} baru + {len(all_previous) - len(detailed_products)} dari sesi sebelumnya)")
        else:
            products_for_excel = detailed_products
        excel_path = save_report(products_for_excel, keyword)
        excel_fname = os.path.basename(excel_path) if excel_path else ""

        # Map untuk tampilan web & upload backend
        session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        mapped = []
        for p in detailed_products:
            price_raw = p.get("price", "")
            digits = re.sub(r'[^0-9]', '', str(price_raw))
            harga_int = int(digits) if digits else 0
            try:
                rating_f = float(str(p.get("rating", 0) or 0).replace(",", "."))
            except Exception:
                rating_f = 0.0
            mapped.append({
                "nama_produk":  p.get("title", "N/A"),
                "harga":        harga_int,
                "platform":     "Tokopedia",
                "rating":       rating_f,
                "terjual":      p.get("sold", "N/A"),
                "url_produk":   p.get("link", ""),
                "gambar_url":   p.get("img", ""),
                "waktu_scrape": p.get("scraped_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                "session_id":   session_id,
                "screenshot":   p.get("screenshot", ""),
                "harga_tinggi": p.get("harga_tinggi", "TIDAK"),
            })

        job["results"]    = mapped
        job["total"]      = len(mapped)
        job["file_excel"] = excel_fname
        job["excel_path"] = excel_path or ""

        upload_status = ""
        if job["backend_url"] and mapped:
            job["message"] = f"Mengirim {len(mapped)} produk ke server..."
            upload_status = _upload_results(
                job["backend_url"], mapped, keyword, session_id,
                job["username"], job["harga_threshold"]
            )
        elif not mapped:
            upload_status = "skip_empty"
        else:
            upload_status = "skip_no_backend"

        job["upload_status"] = upload_status
        job["status"]        = "done"
        job["message"]       = (
            f"✅ Selesai! {len(mapped)} produk ditemukan."
            + (f" | Excel: {excel_fname}" if excel_fname else "")
            + (" | Data terkirim ke server." if upload_status == "ok" else "")
            + (" | ⚠️ Gagal kirim ke server — data tersimpan di Excel lokal." if upload_status == "error" else "")
        )
        job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    except Exception as e:
        job["status"]      = "error"
        job["message"]     = f"Error: {str(e)}"
        job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[Agent Error] {e}")
        import traceback; traceback.print_exc()


def _create_excel(results: list, keyword: str, session_id: str, harga_threshold: int = 350000) -> str:
    """Buat file Excel dari hasil scraping."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        os.makedirs("output", exist_ok=True)
        tanggal       = datetime.now().strftime("%Y-%m-%d")
        keyword_clean = re.sub(r'[\\/*?"<>| ]', "_", keyword)
        filename      = f"hasil_{keyword_clean}_{tanggal}_{session_id}.xlsx"
        filepath      = os.path.join("output", filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Scraping"

        ws.merge_cells("A1:J1")
        ws["A1"] = "SiPantau — Hasil Scraping Tokopedia"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1B4332")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:J2")
        ws["A2"] = f"Keyword: '{keyword}' | Tanggal: {tanggal}"
        ws["A2"].font = Font(bold=True, size=11, color="1B4332")
        ws["A2"].fill = PatternFill("solid", fgColor="D8F3DC")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

        headers = ["No", "Nama Produk", "Harga (Rp)", "Platform", "Rating", "Terjual", "URL Produk", "Waktu Scrape", "Folder Screenshot", "File Screenshot"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="2D6A4F")
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(results):
            row  = 5 + i
            is_expensive = r.get("harga", 0) >= harga_threshold
            
            if is_expensive:
                fill = PatternFill("solid", fgColor="FFCCCC") # Merah pucat
            else:
                fill = PatternFill("solid", fgColor="F0F7F4" if i % 2 == 0 else "FFFFFF")
                
            ss_folder = r.get("ss_folder", "")
            ss_file   = r.get("ss_filename", r.get("screenshot_path", ""))
            data = [i+1, r.get("nama_produk",""), r.get("harga",0), r.get("platform",""),
                    r.get("rating",0), r.get("terjual",""), r.get("url_produk",""),
                    r.get("waktu_scrape",""), ss_folder, ss_file]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.font = Font(size=9, color="990000" if is_expensive else "000000")
                if col == 3:
                    cell.number_format = "#,##0"

        col_widths = [6, 45, 18, 15, 10, 12, 50, 22, 15, 55]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A5"
        wb.save(filepath)
        return filename
    except Exception as e:
        print(f"[Agent] Gagal buat Excel: {e}")
        return ""


def _upload_results(backend_url: str, results: list, keyword: str, session_id: str, username: str, harga_threshold: int = 350000) -> str:
    """Upload hasil scraping ke backend SiPantau. Return 'ok' | 'error'."""
    MAX_RETRY = 3
    for attempt in range(1, MAX_RETRY + 1):
        try:
            url = backend_url.rstrip("/") + "/api/scrape/results"
            payload = {
                "session_id":     session_id,
                "keyword":        keyword,
                "username":       username,
                "platforms":      ["tokopedia"],
                "results":        results,
                "harga_threshold": harga_threshold,
            }
            resp = requests.post(url, json=payload, timeout=30)
            if resp.status_code == 200:
                print(f"[Agent] Hasil dikirim ke backend ({len(results)} produk)")
                return "ok"
            else:
                print(f"[Agent] Backend error (attempt {attempt}): {resp.status_code} — {resp.text[:200]}")
        except Exception as e:
            print(f"[Agent] Gagal kirim ke backend (attempt {attempt}): {e}")
        if attempt < MAX_RETRY:
            time.sleep(3)
    return "error"


# ══════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════
if __name__ == "__main__":
    print()
    print("═"*55)
    print("  SiPantau Agent v" + AGENT_VERSION)
    print("═"*55)
    print("  Jangan tutup jendela ini selama menggunakan SiPantau.")
    print("  Aplikasi SiPantau akan terbuka otomatis di browser.")
    print("═"*55)
    print()

    uvicorn.run(
        app,
        host="127.0.0.1",
        port=AGENT_PORT,
        log_level="warning",
    )
