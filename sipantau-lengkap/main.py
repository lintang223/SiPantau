"""
SiPantau — Sistem Riset Informasi Market
Backend FastAPI + Auth PostgreSQL + Hierarchical RBAC + JWT
Kementerian Lingkungan Hidup dan Kehutanan RI
"""

import os
import time
import hashlib
import logging
from collections import defaultdict
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from typing import List, Optional

import pandas as pd
import psycopg2
import psycopg2.extras
from psycopg2 import pool as pg_pool

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
import bcrypt
from pydantic import BaseModel
from passlib.context import CryptContext

# ── Load .env ─────────────────────────────────────────────────────────────────
load_dotenv()

# ── Config dari environment ───────────────────────────────────────────────────
DB_CONFIG = {
    "host":   os.getenv("DB_HOST", "localhost"),
    "port":   int(os.getenv("DB_PORT", "5050")),
    "dbname": os.getenv("DB_NAME", "sipantau"),
    "user":   os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", "bola"),
}
JWT_SECRET      = os.getenv("JWT_SECRET", "sipantau-dev-secret-GANTI-DI-PRODUKSI")
JWT_ALGORITHM   = "HS256"
JWT_EXPIRE_HRS  = int(os.getenv("JWT_EXPIRE_HOURS", "12"))
ALLOWED_ORIGINS = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",")]

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("sipantau")

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="SiPantau API", version="2.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
os.makedirs("exports", exist_ok=True)

@app.on_event("startup")
def on_startup():
    init_db()

@app.on_event("shutdown")
def on_shutdown():
    global _pool
    if _pool and not _pool.closed:
        _pool.closeall()
        logger.info("Connection pool PostgreSQL ditutup.")

# ── Security Headers ──────────────────────────────────────────────────────────
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"]        = "DENY"
    response.headers["X-XSS-Protection"]       = "1; mode=block"
    return response

# ── Global Exception Handler ──────────────────────────────────────────────────
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled error on {request.url}: {exc}", exc_info=True)
    return JSONResponse(status_code=500, content={"detail": "Terjadi kesalahan internal server."})

# ── Rate Limiting ─────────────────────────────────────────────────────────────
# Hanya catat percobaan LOGIN GAGAL. Dihapus saat berhasil.
def get_lockout_remaining(ip: str, max_attempts: int = 5, window: int = 300) -> int:
    """Kembalikan sisa detik blokir (0 jika tidak diblokir). Menggunakan database."""
    cutoff = (datetime.now() - timedelta(seconds=window)).strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT attempted_at FROM login_logs WHERE ip_address = %s AND status = 'failed' AND attempted_at >= %s ORDER BY attempted_at DESC",
            (ip, cutoff)
        )
        rows = cur.fetchall()
        cur.close()
    
    if len(rows) >= max_attempts:
        first_attempt = datetime.strptime(rows[-1][0], "%Y-%m-%d %H:%M:%S")
        elapsed = (datetime.now() - first_attempt).total_seconds()
        remaining = int(window - elapsed)
        return max(0, remaining)
    return 0

def check_rate_limit(ip: str, max_attempts: int = 5, window: int = 300):
    """Cek apakah IP sedang diblokir karena terlalu banyak gagal login."""
    remaining = get_lockout_remaining(ip, max_attempts, window)
    if remaining > 0:
        raise HTTPException(
            status_code=429,
            detail=f"Terlalu banyak percobaan login. Coba lagi dalam {remaining} detik."
        )

def record_failed_attempt(ip: str):
    """Catatan: Merekam langsung via log_login ke DB, fungsi ini dipertahankan agar kompatibel."""
    pass

def clear_attempts(ip: str):
    """Tandai log gagal sebelumnya sebagai cleared saat login berhasil."""
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE login_logs SET status = 'failed_cleared' WHERE ip_address = %s AND status = 'failed'", (ip,))
        conn.commit()
        cur.close()

def validate_input(value: str, field_name: str, max_length: int = 100) -> str:
    if not value or not value.strip():
        raise HTTPException(status_code=400, detail=f"{field_name} tidak boleh kosong")
    if len(value) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name} terlalu panjang (maks {max_length} karakter)")
    for d in ["'", '"', ";", "--", "/*", "*/", "xp_", "exec", "drop", "truncate"]:
        if d.lower() in value.lower():
            raise HTTPException(status_code=400, detail=f"{field_name} mengandung karakter tidak valid")
    return value.strip()

# ── Password Hashing (bcrypt) ─────────────────────────────────────────────────
def hash_pw(plain: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(plain.encode('utf-8'), salt).decode('utf-8')

def verify_pw(plain: str, hashed: str) -> bool:
    """Verifikasi password. Mendukung bcrypt (baru) dan SHA-256 (legacy)."""
    try:
        return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        # Fallback: cek SHA-256 lama agar admin lama tetap bisa login
        sha = hashlib.sha256(plain.encode('utf-8')).hexdigest()
        return sha == hashed

# ── JWT ───────────────────────────────────────────────────────────────────────
security = HTTPBearer(auto_error=False)

def create_token(username: str, divisi: str, level: int) -> str:
    payload = {
        "sub":    username,
        "divisi": divisi,
        "level":  level,
        "exp":    datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HRS),
        "iat":    datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def get_current_user(request: Request, credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> dict:
    token = None
    if credentials:
        token = credentials.credentials
    if not token:
        token = request.cookies.get("sipantau_token")
        
    if not token:
        raise HTTPException(status_code=401, detail="Token tidak valid atau sudah kadaluarsa. Silakan login ulang.")
        
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if not payload.get("sub"):
            raise HTTPException(status_code=401, detail="Token tidak valid")
        return payload
    except JWTError:
        raise HTTPException(status_code=401, detail="Token tidak valid atau sudah kadaluarsa. Silakan login ulang.")

def require_admin(current_user: dict = Depends(get_current_user)) -> dict:
    """Hanya superadmin dan sekdit yang bisa akses."""
    if current_user.get("level", 99) > 2:
        raise HTTPException(status_code=403, detail="Akses ditolak. Hanya admin yang diizinkan.")
    return current_user

def require_superadmin(current_user: dict = Depends(get_current_user)) -> dict:
    """Hanya superadmin yang bisa akses."""
    if current_user.get("level", 99) > 1:
        raise HTTPException(status_code=403, detail="Akses ditolak. Hanya superadmin yang diizinkan.")
    return current_user

# ── Connection Pool ───────────────────────────────────────────────────────────
_pool: Optional[pg_pool.ThreadedConnectionPool] = None

def get_pool() -> pg_pool.ThreadedConnectionPool:
    global _pool
    if _pool is None or _pool.closed:
        _pool = pg_pool.ThreadedConnectionPool(2, 10, **DB_CONFIG)
        logger.info("Connection pool PostgreSQL dibuat.")
    return _pool

@contextmanager
def get_conn():
    conn = get_pool().getconn()
    try:
        yield conn
    except Exception:
        conn.rollback()
        raise
    finally:
        get_pool().putconn(conn)

# ── Hierarchical RBAC ─────────────────────────────────────────────────────────
DIVISI_LEVEL = {
    "sekditjen": 1,
    "dit_ppsa":  2,
    "balai_gakkum": 3,
}
DIVISI_COLOR = {
    "sekditjen": "#7c3aed",
    "dit_ppsa":  "#0d9488",
    "balai_gakkum": "#2563eb",
}
DEFAULT_ACCESS = [
    ("sekditjen", "dit_ppsa"),
    ("sekditjen", "balai_gakkum"),
    ("dit_ppsa",  "balai_gakkum"),
]

def get_accessible_divisi(conn, user_divisi: str) -> List[str]:
    lvl = DIVISI_LEVEL.get(user_divisi, 99)
    if lvl == 1:
        return list(DIVISI_LEVEL.keys())
    cur = conn.cursor()
    cur.execute(
        "SELECT divisi_target FROM divisi_access WHERE divisi_asal = %s AND can_view = true",
        (user_divisi,)
    )
    result = [r[0] for r in cur.fetchall()]
    cur.close()
    if lvl <= 2 and user_divisi not in result and user_divisi in DIVISI_LEVEL:
        result.insert(0, user_divisi)
    return result

def log_user_activity(conn, username: str, aktivitas: str, detail: str = "", ip_address: str = "unknown"):
    cur = conn.cursor()
    waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """INSERT INTO user_activity (username, aktivitas, detail, ip_address, waktu)
           VALUES (%s, %s, %s, %s, %s)""",
        (username, aktivitas, detail, ip_address, waktu)
    )
    cur.close()

def log_login(conn, username: str, ip: str, user_agent: str, status: str, detail: str = ""):
    """Catat setiap percobaan login ke tabel login_logs."""
    cur = conn.cursor()
    attempted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """INSERT INTO login_logs (username, ip_address, user_agent, status, detail, attempted_at)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        (username, ip, user_agent[:500] if user_agent else "", status, detail, attempted_at)
    )
    cur.close()

def validate_password_complexity(password: str, username: str = "") -> str:
    """Validasi kompleksitas password. Return pesan error, kosong jika valid."""
    if len(password) < 8:
        return "Password minimal 8 karakter"
    has_letter = any(c.isalpha() for c in password)
    has_digit  = any(c.isdigit() for c in password)
    if not has_letter:
        return "Password harus mengandung minimal 1 huruf"
    if not has_digit:
        return "Password harus mengandung minimal 1 angka"
    if username and password.lower() == username.lower():
        return "Password tidak boleh sama dengan username"
    return ""


# ── Init DB ───────────────────────────────────────────────────────────────────
def init_db():
    with get_conn() as conn:
        cur = conn.cursor()

        cur.execute("""CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY, username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, nama TEXT,
            created_at TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS hasil_scraping (
            id SERIAL PRIMARY KEY, session_id TEXT, username TEXT, keyword TEXT,
            nama_produk TEXT, harga BIGINT, platform TEXT, rating REAL,
            terjual TEXT, url_produk TEXT, gambar_url TEXT, waktu_scrape TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS riwayat_session (
            id SERIAL PRIMARY KEY, session_id TEXT, username TEXT, keyword TEXT,
            platforms TEXT, jumlah_data INTEGER, status TEXT, file_excel TEXT, waktu TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS divisi_access (
            id SERIAL PRIMARY KEY,
            divisi_asal   TEXT NOT NULL,
            divisi_target TEXT NOT NULL,
            can_view      BOOLEAN DEFAULT true,
            UNIQUE(divisi_asal, divisi_target)
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS user_activity (
            id SERIAL PRIMARY KEY, username TEXT NOT NULL,
            aktivitas TEXT NOT NULL, detail TEXT,
            ip_address TEXT, waktu TEXT NOT NULL
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS login_logs (
            id           SERIAL PRIMARY KEY,
            username     TEXT,
            ip_address   TEXT,
            user_agent   TEXT,
            status       TEXT NOT NULL,
            detail       TEXT,
            attempted_at TEXT NOT NULL
        )""")
        cur.execute(
            "SELECT COUNT(*) FROM pg_indexes "
            "WHERE tablename='login_logs' AND indexname='idx_login_logs_attempted_at'"
        )
        if cur.fetchone()[0] == 0:
            cur.execute("CREATE INDEX idx_login_logs_attempted_at ON login_logs(attempted_at)")

        # Column migrations
        migrations = [
            ("users",           "divisi",           "TEXT DEFAULT 'balai_gakkum'"),
            ("users",           "level",            "INTEGER DEFAULT 3"),
            ("users",           "can_export",       "BOOLEAN DEFAULT true"),
            ("users",           "can_manage_users", "BOOLEAN DEFAULT false"),
            ("users",           "foto_profil",      "TEXT"),
            ("users",           "updated_at",       "TEXT"),
            ("users",           "deleted_at",       "TEXT"),
            ("hasil_scraping",  "username",         "TEXT"),
            ("riwayat_session", "username",         "TEXT"),
            ("riwayat_session", "divisi",           "TEXT DEFAULT 'balai_gakkum'"),
        ]
        for table, col, col_type in migrations:
            cur.execute(
                "SELECT COUNT(*) FROM information_schema.columns WHERE table_name=%s AND column_name=%s",
                (table, col)
            )
            if cur.fetchone()[0] == 0:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}")

        # Hapus kolom password_plain jika masih ada (cleanup legacy)
        cur.execute(
            "SELECT COUNT(*) FROM information_schema.columns WHERE table_name='users' AND column_name='password_plain'"
        )
        if cur.fetchone()[0] > 0:
            cur.execute("ALTER TABLE users DROP COLUMN password_plain")
            logger.info("Kolom password_plain dihapus dari tabel users.")

        # Seed access rules
        for asal, target in DEFAULT_ACCESS:
            cur.execute(
                "INSERT INTO divisi_access (divisi_asal, divisi_target) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                (asal, target)
            )

        # Default admin — password dari env var, TIDAK hardcoded di source code
        cur.execute("SELECT COUNT(*) FROM users WHERE username='admin'")
        if cur.fetchone()[0] == 0:
            import secrets, string
            _admin_pw = os.getenv("ADMIN_DEFAULT_PASSWORD", "")
            if not _admin_pw:
                # Tidak ada di .env — generate password acak yang aman
                _chars    = string.ascii_letters + string.digits + "!@#$%"
                _admin_pw = "".join(secrets.choice(_chars) for _ in range(16))
                logger.warning("="*60)
                logger.warning("ADMIN DEFAULT PASSWORD (hanya tampil sekali):")
                logger.warning(f"  username : admin")
                logger.warning(f"  password : {_admin_pw}")
                logger.warning("Salin ke .env: ADMIN_DEFAULT_PASSWORD=<password>")
                logger.warning("="*60)
            cur.execute(
                """INSERT INTO users (username,password,nama,divisi,level,can_export,can_manage_users,created_at)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                ("admin", hash_pw(_admin_pw), "Administrator",
                 "sekditjen", 1, True, True, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            )

        # Sync level
        cur.execute("SELECT username, divisi FROM users WHERE level IS NULL OR level = 0")
        for uname, divisi in cur.fetchall():
            lvl = DIVISI_LEVEL.get(divisi or "balai_gakkum", 3)
            cur.execute("UPDATE users SET level=%s WHERE username=%s", (lvl, uname))

        conn.commit()
        cur.close()
    print("Database PostgreSQL siap!")

# ── Models ────────────────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

class ChangePasswordRequest(BaseModel):
    username: str
    password_lama: str
    password_baru: str

class TambahUserRequest(BaseModel):
    username: str
    password: str
    nama: str
    divisi: str = "balai_gakkum"

class ResetPasswordRequest(BaseModel):
    username: str
    password_baru: str

class UpdateProfilRequest(BaseModel):
    nama: str

class UpdateFotoRequest(BaseModel):
    foto: str

class ScrapeRequest(BaseModel):
    keyword: str
    platforms: List[str]
    max_pages: int = 3
    max_load_more: int = 5
    harga_threshold: int = 350000
    min_price: Optional[int] = 0
    max_price: Optional[int] = 999999999
    sort_by: Optional[str] = "relevance"
    username: Optional[str] = None

class ScrapeResultsRequest(BaseModel):
    session_id: str
    keyword: str
    username: str
    platforms: List[str]
    results: List[dict]
    harga_threshold: int = 350000

class ExportRequest(BaseModel):
    session_id: str
    keyword: str

# ── Health ────────────────────────────────────────────────────────────────────
@app.get("/health")
def health_check():
    try:
        with get_conn() as conn:
            conn.cursor().execute("SELECT 1")
        return {"status": "ok", "app": "SiPantau", "versi": "2.0.0", "db": "PostgreSQL"}
    except Exception as e:
        logger.error(f"Health check DB error: {e}")
        raise HTTPException(status_code=500, detail=f"DB error: {str(e)}")

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.get("/api/auth/lockout-status")
def lockout_status(request: Request):
    """Cek sisa waktu blokir untuk IP pemanggil. Frontend pakai ini saat halaman dimuat."""
    ip = request.client.host if request.client else "unknown"
    remaining = get_lockout_remaining(ip)
    return {"locked": remaining > 0, "remaining_seconds": remaining}

@app.post("/api/auth/login")
def login(req: LoginRequest, request: Request, response: Response):
    ip         = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("User-Agent", "")

    # Cek blokir SEBELUM apapun
    try:
        check_rate_limit(ip)
    except HTTPException as e:
        with get_conn() as conn:
            log_login(conn, req.username, ip, user_agent, "blocked", "IP diblokir karena terlalu banyak percobaan gagal")
            conn.commit()
        raise e

    username_clean = req.username.strip().lstrip('@')
    
    validate_input(username_clean, "Username", max_length=50)
    if not (1 <= len(req.password) <= 100):
        raise HTTPException(status_code=400, detail="Password tidak valid")

    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM users WHERE username = %s AND deleted_at IS NULL", (username_clean,))
        user = cur.fetchone()
        cur.close()

    if not user or not verify_pw(req.password, user["password"]):
        # record_failed_attempt is obsolete, we query DB
        remaining_attempts = 5 - (5 - get_lockout_remaining(ip)) # Rough calculation, just use remaining logic
        detail_msg = "Username tidak ditemukan atau dinonaktifkan" if not user else "Password salah."
        logger.warning(f"Login gagal untuk username '{username_clean}' dari IP {ip}")
        with get_conn() as conn:
            log_login(conn, username_clean, ip, user_agent, "failed", detail_msg)
            conn.commit()
        # Hitung sisa setelah insert
        remaining_seconds = get_lockout_remaining(ip)
        if remaining_seconds > 0:
            raise HTTPException(status_code=429, detail=f"Terlalu banyak percobaan login. Coba lagi dalam {remaining_seconds} detik.")
        raise HTTPException(status_code=401, detail=detail_msg)

    # Login berhasil
    clear_attempts(ip)
    divisi = user.get("divisi") or "balai_gakkum"
    level  = user.get("level") or DIVISI_LEVEL.get(divisi, 3)
    token  = create_token(user["username"], divisi, level)

    with get_conn() as conn:
        accessible = get_accessible_divisi(conn, divisi)
        log_user_activity(conn, username_clean, "Login", "User berhasil login", ip)
        log_login(conn, username_clean, ip, user_agent, "success", "Login berhasil")
        conn.commit()

    logger.info(f"Login berhasil: {username_clean} dari IP {ip}")
    
    # Set HttpOnly Cookie — token TIDAK dikirim di response body (cegah akses dari JS)
    response.set_cookie(
        key="sipantau_token",
        value=token,
        httponly=True,
        max_age=JWT_EXPIRE_HRS * 3600,
        samesite="lax", # Ganti ke "none" jika beda domain (wajib secure=True)
        secure=False,   # Ubah ke True jika sudah pakai HTTPS
    )

    return {
        "success": True,
        # Token SENGAJA tidak dikembalikan di sini.
        # Auth sepenuhnya via HttpOnly Cookie yang tidak bisa dibaca JS.
        "user": {
            "username":          user["username"],
            "nama":              user["nama"],
            "divisi":            divisi,
            "level":             level,
            "can_export":        bool(user.get("can_export", True)),
            "can_manage_users":  bool(user.get("can_manage_users", False)),
            "accessible_divisi": accessible,
            "divisi_color":      DIVISI_COLOR.get(divisi, "#374151"),
        }
    }

@app.post("/api/auth/logout")
def logout(response: Response):
    """Menghapus cookie sipantau_token dari browser."""
    response.delete_cookie(key="sipantau_token", path="/", samesite="lax")
    return {"success": True, "message": "Berhasil logout"}

@app.get("/api/auth/me")
def get_me(current_user: dict = Depends(get_current_user)):
    username = current_user["sub"]
    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
        if not user:
            cur.close()
            raise HTTPException(status_code=404, detail="User tidak ditemukan")
        divisi     = user.get("divisi") or "balai_gakkum"
        level      = user.get("level") or DIVISI_LEVEL.get(divisi, 3)
        accessible = get_accessible_divisi(conn, divisi)
        cur.close()
    return {
        "success": True,
        "user": {
            "username":          user["username"],
            "nama":              user["nama"],
            "divisi":            divisi,
            "level":             level,
            "can_export":        bool(user.get("can_export", True)),
            "can_manage_users":  bool(user.get("can_manage_users", False)),
            "accessible_divisi": accessible,
            "divisi_color":      DIVISI_COLOR.get(divisi, "#374151"),
            "foto_profil":       user.get("foto_profil"),
        }
    }

@app.post("/api/auth/ganti-password")
def ganti_password(req: ChangePasswordRequest, request: Request, current_user: dict = Depends(get_current_user)):
    ip = request.client.host if request.client else "unknown"
    if current_user["sub"] != req.username:
        raise HTTPException(status_code=403, detail="Tidak bisa ganti password user lain")
    err = validate_password_complexity(req.password_baru, req.username)
    if err:
        raise HTTPException(status_code=400, detail=err)

    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM users WHERE username = %s", (req.username,))
        user = cur.fetchone()
        if not user or not verify_pw(req.password_lama, user["password"]):
            cur.close()
            raise HTTPException(status_code=401, detail="Password lama salah")
        cur.execute(
            "UPDATE users SET password=%s WHERE username=%s",
            (hash_pw(req.password_baru), req.username)
        )
        log_user_activity(conn, req.username, "Ganti Password", "User mengubah password miliknya", ip)
        conn.commit()
        cur.close()
    return {"success": True, "message": "Password berhasil diubah"}

@app.put("/api/auth/update-profil")
def update_profil(req: UpdateProfilRequest, request: Request, current_user: dict = Depends(get_current_user)):
    username = current_user["sub"]
    ip = request.client.host if request.client else "unknown"
    nama_bersih = req.nama.strip()
    if not nama_bersih:
        raise HTTPException(status_code=400, detail="Nama tidak boleh kosong")
    if len(nama_bersih) > 100:
        raise HTTPException(status_code=400, detail="Nama terlalu panjang (maks 100 karakter)")

    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("UPDATE users SET nama=%s WHERE username=%s", (nama_bersih, username))
        if cur.rowcount == 0:
            cur.close()
            raise HTTPException(status_code=404, detail="User tidak ditemukan")
        # Ambil data user terbaru
        cur.execute("SELECT * FROM users WHERE username=%s", (username,))
        user = cur.fetchone()
        divisi = user.get("divisi") or "balai_gakkum"
        level  = user.get("level") or DIVISI_LEVEL.get(divisi, 3)
        accessible = get_accessible_divisi(conn, divisi)
        log_user_activity(conn, username, "Update Profil", f"Nama diubah menjadi '{nama_bersih}'", ip)
        conn.commit()
        cur.close()

    return {
        "success": True,
        "message": "Profil berhasil diperbarui",
        "user": {
            "username":          user["username"],
            "nama":              user["nama"],
            "divisi":            divisi,
            "level":             level,
            "can_export":        bool(user.get("can_export", True)),
            "can_manage_users":  bool(user.get("can_manage_users", False)),
            "accessible_divisi": accessible,
            "divisi_color":      DIVISI_COLOR.get(divisi, "#374151"),
            "foto_profil":       user.get("foto_profil"),
        }
    }

@app.put("/api/auth/update-foto")
def update_foto(req: UpdateFotoRequest, request: Request, current_user: dict = Depends(get_current_user)):
    username = current_user["sub"]
    ip = request.client.host if request.client else "unknown"

    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("UPDATE users SET foto_profil=%s WHERE username=%s", (req.foto, username))
        if cur.rowcount == 0:
            cur.close()
            raise HTTPException(status_code=404, detail="User tidak ditemukan")
        
        cur.execute("SELECT * FROM users WHERE username=%s", (username,))
        user = cur.fetchone()
        divisi = user.get("divisi") or "balai_gakkum"
        level  = user.get("level") or DIVISI_LEVEL.get(divisi, 3)
        accessible = get_accessible_divisi(conn, divisi)
        
        log_user_activity(conn, username, "Update Foto Profil", "User mengubah foto profil", ip)
        conn.commit()
        cur.close()

    return {
        "success": True,
        "message": "Foto profil berhasil diperbarui",
        "user": {
            "username":          user["username"],
            "nama":              user["nama"],
            "divisi":            divisi,
            "level":             level,
            "can_export":        bool(user.get("can_export", True)),
            "can_manage_users":  bool(user.get("can_manage_users", False)),
            "accessible_divisi": accessible,
            "divisi_color":      DIVISI_COLOR.get(divisi, "#374151"),
            "foto_profil":       user.get("foto_profil"),
        }
    }

@app.get("/api/user-activity")
def get_user_activity(current_user: dict = Depends(require_admin)):
    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM user_activity ORDER BY id DESC LIMIT 500")
        rows = cur.fetchall()
        cur.close()
    return {"activity": [dict(r) for r in rows]}

@app.get("/api/login-logs")
def get_login_logs(
    status: str = "",
    username: str = "",
    limit: int = 500,
    current_user: dict = Depends(require_superadmin)
):
    """Ambil log percobaan login. Hanya superadmin (sekditjen) yang bisa akses."""
    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        conditions = []
        params     = []
        if status:
            conditions.append("status = %s")
            params.append(status)
        if username:
            conditions.append("username ILIKE %s")
            params.append(f"%{username}%")
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params.append(min(limit, 1000))
        cur.execute(f"SELECT * FROM login_logs {where} ORDER BY id DESC LIMIT %s", params)
        rows = cur.fetchall()
        cur.close()
    return {"logs": [dict(r) for r in rows]}

# ── User Management ───────────────────────────────────────────────────────────
@app.get("/api/users")
def get_users(
    include_deleted: bool = False,
    current_user: dict = Depends(require_admin)
):
    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        where = "" if (include_deleted and current_user.get("level", 99) == 1) else "WHERE deleted_at IS NULL"
        cur.execute(
            f"SELECT id,username,nama,divisi,level,can_export,can_manage_users,created_at,updated_at,deleted_at FROM users {where} ORDER BY id"
        )
        rows = cur.fetchall()
        cur.close()
    return {"users": [dict(r) for r in rows]}

@app.post("/api/users")
def tambah_user(req: TambahUserRequest, current_user: dict = Depends(require_admin)):
    validate_input(req.username, "Username", max_length=50)
    validate_input(req.nama, "Nama", max_length=100)
    if req.divisi not in DIVISI_LEVEL:
        raise HTTPException(status_code=400, detail=f"Divisi tidak valid. Pilih: {list(DIVISI_LEVEL.keys())}")
    pw_err = validate_password_complexity(req.password, req.username)
    if pw_err:
        raise HTTPException(status_code=400, detail=pw_err)

    level      = DIVISI_LEVEL.get(req.divisi, 3)
    can_manage = req.divisi in ("sekditjen",)
    now        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_conn() as conn:
        cur = conn.cursor()
        try:
            # Cek apakah username sudah ada tapi di-soft-delete — restore saja
            cur.execute("SELECT id, deleted_at FROM users WHERE username = %s", (req.username,))
            existing = cur.fetchone()
            if existing and existing[1] is not None:
                # User lama yang di-soft-delete — update dengan data baru
                cur.execute(
                    """UPDATE users SET password=%s, nama=%s, divisi=%s, level=%s,
                       can_export=%s, can_manage_users=%s, updated_at=%s, deleted_at=NULL
                       WHERE username=%s""",
                    (hash_pw(req.password), req.nama, req.divisi, level, True, can_manage, now, req.username)
                )
            elif existing:
                raise HTTPException(status_code=400, detail="Username sudah digunakan")
            else:
                cur.execute(
                    """INSERT INTO users (username,password,nama,divisi,level,can_export,can_manage_users,created_at)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (req.username, hash_pw(req.password), req.nama,
                     req.divisi, level, True, can_manage, now)
                )
            conn.commit()
        except HTTPException:
            conn.rollback(); raise
        except psycopg2.errors.UniqueViolation:
            conn.rollback()
            raise HTTPException(status_code=400, detail="Username sudah digunakan")
        finally:
            cur.close()
    logger.info(f"User baru dibuat: {req.username} (divisi: {req.divisi}) oleh {current_user['sub']}")
    return {"success": True, "message": f"User '{req.username}' berhasil ditambahkan"}

@app.delete("/api/users/{username}")
def hapus_user(username: str, current_user: dict = Depends(require_admin)):
    if username == "admin":
        raise HTTPException(status_code=400, detail="Akun admin utama tidak bisa dihapus")
    if username == current_user["sub"]:
        raise HTTPException(status_code=400, detail="Tidak bisa menghapus akun sendiri")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.cursor()
        # Soft delete — set deleted_at, TIDAK menghapus row dari database
        cur.execute(
            "UPDATE users SET deleted_at=%s, updated_at=%s WHERE username=%s AND deleted_at IS NULL",
            (now, now, username)
        )
        if cur.rowcount == 0:
            cur.close()
            raise HTTPException(status_code=404, detail="User tidak ditemukan atau sudah dinonaktifkan")
        conn.commit()
        cur.close()
    logger.info(f"User dinonaktifkan (soft delete): {username} oleh {current_user['sub']}")
    return {"success": True, "message": f"User '{username}' berhasil dinonaktifkan"}

@app.put("/api/users/{username}/restore")
def restore_user(username: str, current_user: dict = Depends(require_superadmin)):
    """Pulihkan user yang di-soft-delete. Hanya superadmin."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE users SET deleted_at=NULL, updated_at=%s WHERE username=%s AND deleted_at IS NOT NULL",
            (now, username)
        )
        if cur.rowcount == 0:
            cur.close()
            raise HTTPException(status_code=404, detail="User tidak ditemukan atau masih aktif")
        conn.commit()
        cur.close()
    logger.info(f"User dipulihkan: {username} oleh {current_user['sub']}")
    return {"success": True, "message": f"User '{username}' berhasil dipulihkan"}

@app.post("/api/users/reset-password")
def reset_password_user(req: ResetPasswordRequest, current_user: dict = Depends(require_admin)):
    pw_err = validate_password_complexity(req.password_baru, req.username)
    if pw_err:
        raise HTTPException(status_code=400, detail=pw_err)

    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM users WHERE username=%s", (req.username,))
        if cur.fetchone()[0] == 0:
            cur.close()
            raise HTTPException(status_code=404, detail="User tidak ditemukan")
        cur.execute(
            "UPDATE users SET password=%s WHERE username=%s",
            (hash_pw(req.password_baru), req.username)
        )
        conn.commit()
        cur.close()
    logger.info(f"Password direset untuk: {req.username} oleh {current_user['sub']}")
    return {"success": True, "message": f"Password '{req.username}' berhasil direset"}

# ── Riwayat ───────────────────────────────────────────────────────────────────
@app.get("/api/riwayat")
def get_riwayat(
    username: str = "", divisi: str = "", view_all: bool = False,
    current_user: dict = Depends(get_current_user)
):
    with get_conn() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if view_all:
            if current_user.get("level", 99) > 2:
                raise HTTPException(status_code=403, detail="Akses ditolak")
            cur.execute(
                "SELECT rs.*, u.divisi as user_divisi FROM riwayat_session rs LEFT JOIN users u ON rs.username=u.username ORDER BY rs.id DESC LIMIT 200"
            )
        elif divisi:
            user_divisi = current_user.get("divisi", "balai_gakkum")
            accessible = get_accessible_divisi(conn, user_divisi)
            if divisi not in accessible:
                raise HTTPException(status_code=403, detail="Akses ditolak")
            cur.execute("SELECT username FROM users WHERE divisi=%s", (divisi,))
            unames = [r["username"] for r in cur.fetchall()]
            if not unames:
                cur.close()
                return {"riwayat": []}
            ph = ",".join(["%s"] * len(unames))
            cur.execute(
                f"SELECT rs.*, u.divisi as user_divisi FROM riwayat_session rs LEFT JOIN users u ON rs.username=u.username WHERE rs.username IN ({ph}) ORDER BY rs.id DESC LIMIT 100",
                unames
            )
        elif username:
            # User hanya bisa lihat riwayat sendiri kecuali admin
            if current_user.get("level", 99) > 2 and username != current_user["sub"]:
                raise HTTPException(status_code=403, detail="Akses ditolak")
            cur.execute(
                "SELECT rs.*, u.divisi as user_divisi FROM riwayat_session rs LEFT JOIN users u ON rs.username=u.username WHERE rs.username=%s ORDER BY rs.id DESC LIMIT 50", (username,)
            )
        else:
            if current_user.get("level", 99) > 2:
                cur.execute(
                    "SELECT rs.*, u.divisi as user_divisi FROM riwayat_session rs LEFT JOIN users u ON rs.username=u.username WHERE rs.username=%s ORDER BY rs.id DESC LIMIT 50",
                    (current_user["sub"],)
                )
            else:
                cur.execute("SELECT rs.*, u.divisi as user_divisi FROM riwayat_session rs LEFT JOIN users u ON rs.username=u.username ORDER BY rs.id DESC LIMIT 50")
        rows = cur.fetchall()
        cur.close()
    return {"riwayat": [dict(r) for r in rows]}

@app.get("/api/riwayat/divisi-list")
def get_divisi_list(current_user: dict = Depends(get_current_user)):
    user_divisi = current_user.get("divisi", "balai_gakkum")
    with get_conn() as conn:
        accessible = get_accessible_divisi(conn, user_divisi)
    return {"divisi_list": accessible}

# ── Stats ─────────────────────────────────────────────────────────────────────
@app.get("/api/stats")
def get_stats(
    username: str = "", divisi: str = "",
    current_user: dict = Depends(get_current_user)
):
    with get_conn() as conn:
        cur = conn.cursor()

        def count(sql, params=()):
            cur.execute(sql, params)
            return cur.fetchone()[0]

        if divisi:
            cur.execute("SELECT username FROM users WHERE divisi=%s", (divisi,))
            unames = tuple(r[0] for r in cur.fetchall())
            if not unames:
                cur.close()
                return {"total": 0, "tokopedia": 0, "ekspor": 0}
            ph = ",".join(["%s"] * len(unames))
            total  = count(f"SELECT COUNT(*) FROM hasil_scraping WHERE username IN ({ph})", unames)
            tokped = count(f"SELECT COUNT(*) FROM hasil_scraping WHERE username IN ({ph}) AND LOWER(platform)='tokopedia'", unames)
            ekspor = count(f"SELECT COUNT(*) FROM riwayat_session WHERE username IN ({ph})", unames)
        elif username:
            total  = count("SELECT COUNT(*) FROM hasil_scraping WHERE username=%s", (username,))
            tokped = count("SELECT COUNT(*) FROM hasil_scraping WHERE username=%s AND LOWER(platform)='tokopedia'", (username,))
            ekspor = count("SELECT COUNT(*) FROM riwayat_session WHERE username=%s", (username,))
        else:
            total  = count("SELECT COUNT(*) FROM hasil_scraping")
            tokped = count("SELECT COUNT(*) FROM hasil_scraping WHERE LOWER(platform)='tokopedia'")
            ekspor = count("SELECT COUNT(*) FROM riwayat_session")
        cur.close()
    return {"total": total, "tokopedia": tokped, "ekspor": ekspor}

# ── Scraping ──────────────────────────────────────────────────────────────────
@app.post("/api/scrape")
async def scrape(req: ScrapeRequest, current_user: dict = Depends(get_current_user)):
    session_id  = datetime.now().strftime("%Y%m%d_%H%M%S")
    username    = current_user["sub"]
    all_results = []
    for platform in req.platforms:
        try:
            results = generate_placeholder(platform, req.keyword, req.max_pages)
            waktu   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for r in results:
                r["waktu_scrape"] = waktu
                r["session_id"]   = session_id
            all_results.extend(results)
        except Exception as e:
            logger.error(f"Error scraping {platform}: {e}")
    save_to_db(all_results, session_id, req.keyword, req.platforms, username)
    filename = export_to_excel_file(all_results, req.keyword, session_id, req.harga_threshold)
    return {
        "session_id": session_id, "keyword": req.keyword,
        "total": len(all_results), "results": all_results, "file_excel": filename
    }

@app.post("/api/scrape/results")
def receive_scrape_results(req: ScrapeResultsRequest, current_user: dict = Depends(get_current_user)):
    save_to_db(req.results, req.session_id, req.keyword, req.platforms, req.username)
    filename = export_to_excel_file(req.results, req.keyword, req.session_id, req.harga_threshold)
    return {"success": True, "message": f"{len(req.results)} data disimpan", "file_excel": filename}

# ── Export ────────────────────────────────────────────────────────────────────
@app.post("/api/export")
def export_excel(req: ExportRequest, current_user: dict = Depends(get_current_user)):
    with get_conn() as conn:
        df = pd.read_sql("SELECT * FROM hasil_scraping WHERE session_id = %s", conn, params=(req.session_id,))
    if df.empty:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    filename = export_to_excel_file(df.to_dict("records"), req.keyword, req.session_id)
    return FileResponse(
        path=f"exports/{filename}", filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/export/download/{filename}")
def download_excel(filename: str, current_user: dict = Depends(get_current_user)):
    # Validasi nama file — cegah path traversal
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Nama file tidak valid")
    filepath = f"exports/{filename}"
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    return FileResponse(
        path=filepath, filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ── Helpers ───────────────────────────────────────────────────────────────────
def save_to_db(results, session_id, keyword, platforms, username="unknown"):
    with get_conn() as conn:
        cur = conn.cursor()
        for r in results:
            cur.execute(
                """INSERT INTO hasil_scraping
                   (session_id,username,keyword,nama_produk,harga,platform,rating,terjual,url_produk,gambar_url,waktu_scrape)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (session_id, username, keyword, r.get("nama_produk",""), r.get("harga",0),
                 r.get("platform",""), r.get("rating",0), r.get("terjual",""),
                 r.get("url_produk",""), r.get("gambar_url",""), r.get("waktu_scrape",""))
            )
        cur.execute(
            """INSERT INTO riwayat_session (session_id,username,keyword,platforms,jumlah_data,status,waktu)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (session_id, username, keyword, ", ".join(platforms), len(results),
             "Selesai", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        cur.close()

def export_to_excel_file(results, keyword, session_id, harga_threshold=350000):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    tanggal  = datetime.now().strftime("%Y-%m-%d")
    filename = f"hasil_scraping_{keyword.replace(' ','_')}_{tanggal}_{session_id}.xlsx"
    filepath = f"exports/{filename}"
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "Data Scraping"

    ws.merge_cells("A1:I1")
    ws["A1"] = "KEMENTERIAN LINGKUNGAN HIDUP DAN KEHUTANAN REPUBLIK INDONESIA"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1B4332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = f"SiPantau — Hasil Scraping: '{keyword}' | Tanggal: {tanggal}"
    ws["A2"].font      = Font(bold=True, size=11, color="1B4332")
    ws["A2"].fill      = PatternFill("solid", fgColor="D8F3DC")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    headers = ["No","Nama Produk","Harga (Rp)","Platform","Rating","Terjual","URL Produk","Waktu Scrape"]
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=4, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="2D6A4F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    for i, r in enumerate(results):
        row    = 5 + i
        is_exp = r.get("harga", 0) >= 1000000
        fill   = PatternFill("solid", fgColor="FFCCCC") if is_exp else PatternFill("solid", fgColor="F0F7F4" if i%2==0 else "FFFFFF")
        data   = [i+1, r.get("nama_produk",""), r.get("harga",0), r.get("platform",""),
                  r.get("rating",0), r.get("terjual",""), r.get("url_produk",""), r.get("waktu_scrape","")]
        for col, val in enumerate(data, 1):
            cell      = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = Font(size=9, color="990000" if is_exp else "000000")
            if col == 3: cell.number_format = "#,##0"; cell.alignment = Alignment(horizontal="right")
            elif col == 5: cell.number_format = "0.0"

    for i, w in enumerate([6,45,18,15,10,12,50,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    wb.save(filepath)
    return filename

def generate_placeholder(platform, keyword, max_pages):
    import random
    produk = [f"Kayu Jati {keyword}", f"Bambu {keyword}", f"Rotan {keyword}",
              f"Madu Hutan {keyword}", f"Gaharu {keyword}", f"Kayu Sengon {keyword}"]
    return [{"nama_produk": random.choice(produk)+f" #{random.randint(1,99)}",
             "harga": random.randint(10000,500000), "platform": platform.capitalize(),
             "rating": round(random.uniform(3.0,5.0),1), "terjual": f"{random.randint(1,500)}rb+",
             "url_produk": f"https://{platform}.co.id/produk/{keyword.replace(' ','-')}",
             "gambar_url": ""} for _ in range(max_pages*10)]

if __name__ == "__main__":
    import uvicorn
    print("SiPantau Backend v2.0 berjalan di http://localhost:8000")
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)