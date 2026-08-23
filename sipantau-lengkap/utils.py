import os
from datetime import datetime, timedelta
from fastapi import HTTPException
from database import get_conn
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def log_user_activity(conn, username: str, aktivitas: str, detail: str = "", ip_address: str = "unknown"):
    cur = conn.cursor()
    waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """INSERT INTO user_activity (username, aktivitas, detail, ip_address, waktu)
           VALUES (%s, %s, %s, %s, %s)""",
        (username, aktivitas, detail, ip_address, waktu)
    )
    cur.close()

def log_login(conn, username: str, ip: str, user_agent: str, status: str, detail: str = ""):
    cur = conn.cursor()
    attempted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """INSERT INTO login_logs (username, ip_address, user_agent, status, detail, attempted_at)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        (username, ip, user_agent[:500] if user_agent else "", status, detail, attempted_at)
    )
    cur.close()

def get_lockout_remaining(ip: str, max_attempts: int = 5, window: int = 300) -> int:
    cutoff = (datetime.now() - timedelta(seconds=window)).strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT attempted_at FROM login_logs WHERE ip_address = %s AND status = 'failed' AND attempted_at >= %s ORDER BY attempted_at DESC",
            (ip, cutoff)
        )
        rows = cur.fetchall()
        cur.close()
    
    if len(rows) >= max_attempts:
        first_attempt = datetime.strptime(rows[-1]["attempted_at"], "%Y-%m-%d %H:%M:%S")
        elapsed = (datetime.now() - first_attempt).total_seconds()
        remaining = int(window - elapsed)
        return max(0, remaining)
    return 0

def check_rate_limit(ip: str, max_attempts: int = 5, window: int = 300):
    remaining = get_lockout_remaining(ip, max_attempts, window)
    if remaining > 0:
        raise HTTPException(
            status_code=429,
            detail=f"Terlalu banyak percobaan login. Coba lagi dalam {remaining} detik."
        )


def clear_attempts(ip: str):
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE login_logs SET status = 'failed_cleared' WHERE ip_address = %s AND status = 'failed'", (ip,))
        conn.commit()

def validate_input(value: str, field_name: str, max_length: int = 100) -> str:
    if not value or not value.strip():
        raise HTTPException(status_code=400, detail=f"{field_name} tidak boleh kosong")
    if len(value) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name} terlalu panjang (maks {max_length} karakter)")
    return value.strip()

def save_to_db(results, session_id, keyword, platforms, username="unknown", file_excel=""):
    with get_conn() as conn:
        cur = conn.cursor()
        
        insert_query = """
            INSERT INTO hasil_scraping
            (session_id,username,keyword,nama_produk,harga,platform,rating,terjual,url_produk,gambar_url,waktu_scrape)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        
        data_to_insert = [
            (
                session_id, username, keyword, r.get("nama_produk", ""), r.get("harga", 0),
                r.get("platform", ""), r.get("rating", 0), r.get("terjual", ""),
                r.get("url_produk", ""), r.get("gambar_url", ""), r.get("waktu_scrape", "")
            ) for r in results
        ]
        
        if data_to_insert:
            cur.executemany(insert_query, data_to_insert)

        cur.execute(
            """INSERT INTO riwayat_session (session_id,username,keyword,platforms,jumlah_data,status,waktu,file_excel)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            (session_id, username, keyword, ", ".join(platforms), len(results),
             "Selesai", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), file_excel)
        )
        conn.commit()
        cur.close()

def export_to_excel_file(results, keyword, session_id, harga_threshold=350000):
    tanggal  = datetime.now().strftime("%Y-%m-%d")
    filename = f"hasil_scraping_{keyword.replace(' ','_')}_{tanggal}_{session_id}.xlsx"
    filepath = f"exports/{filename}"
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "Data Scraping"

    ws.merge_cells("A1:I1")
    ws["A1"] = "KEMENTERIAN KEHUTANAN REPUBLIK INDONESIA"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1B4332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = f"SiPantau — Hasil Scraping: '{keyword}' | Tanggal: {tanggal}"
    ws["A2"].font      = Font(bold=True, size=11, color="1B4332")
    ws["A2"].fill      = PatternFill("solid", fgColor="D8F3DC")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    headers = ["No","Nama Produk","Harga (Rp)","Platform","Rating","Terjual","URL Produk","Waktu Scrape"]
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=4, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="2D6A4F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    for i, r in enumerate(results):
        row    = 5 + i
        is_exp = r.get("harga", 0) >= 1000000
        fill   = PatternFill("solid", fgColor="FFCCCC") if is_exp else PatternFill("solid", fgColor="F0F7F4" if i%2==0 else "FFFFFF")
        data   = [i+1, r.get("nama_produk",""), r.get("harga",0), r.get("platform",""),
                  r.get("rating",0), r.get("terjual",""), r.get("url_produk",""), r.get("waktu_scrape","")]
        for col, val in enumerate(data, 1):
            cell      = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = Font(size=9, color="990000" if is_exp else "000000")
            if col == 3: cell.number_format = "#,##0"; cell.alignment = Alignment(horizontal="right")
            elif col == 5: cell.number_format = "0.0"

    for i, w in enumerate([6,45,18,15,10,12,50,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    wb.save(filepath)
    return filename
