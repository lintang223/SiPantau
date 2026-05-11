"""
SiPantau — Sistem Riset Informasi Market
Backend FastAPI + Auth PostgreSQL
Kementerian Lingkungan Hidup dan Kehutanan RI
"""

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import pandas as pd
import sqlite3
import hashlib
import os

app = FastAPI(title="SiPantau API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

os.makedirs("exports", exist_ok=True)

# =============================================
# SECURITY — Rate limiting sederhana
# =============================================
from collections import defaultdict
import time

login_attempts: dict = defaultdict(list)

def check_rate_limit(ip: str, max_attempts: int = 5, window: int = 60):
    """Blokir IP yang gagal login lebih dari 5x dalam 60 detik"""
    now = time.time()
    attempts = [t for t in login_attempts[ip] if now - t < window]
    login_attempts[ip] = attempts
    if len(attempts) >= max_attempts:
        raise HTTPException(status_code=429, detail="Terlalu banyak percobaan login. Coba lagi dalam 1 menit.")
    login_attempts[ip].append(now)

def validate_input(value: str, field_name: str, max_length: int = 100):
    """Validasi input — cegah karakter berbahaya"""
    if not value or not value.strip():
        raise HTTPException(status_code=400, detail=f"{field_name} tidak boleh kosong")
    if len(value) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name} terlalu panjang (max {max_length} karakter)")
    # Cegah karakter SQL injection meski sudah pakai parameterized query
    dangerous = ["'", '"', ";", "--", "/*", "*/", "xp_", "exec", "drop", "truncate"]
    for d in dangerous:
        if d.lower() in value.lower():
            raise HTTPException(status_code=400, detail=f"{field_name} mengandung karakter tidak valid")
    return value.strip()

@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    return response

DB_FILE = "sipantau.db"

def get_conn():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def hash_pw(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

# =============================================
# INIT DATABASE
# =============================================
def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            username       TEXT UNIQUE NOT NULL,
            password       TEXT NOT NULL,
            password_plain TEXT,
            nama           TEXT,
            role           TEXT DEFAULT 'user',
            created_at     TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS hasil_scraping (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id   TEXT,
            username     TEXT,
            keyword      TEXT,
            nama_produk  TEXT,
            harga        INTEGER,
            platform     TEXT,
            rating       REAL,
            terjual      TEXT,
            url_produk   TEXT,
            gambar_url   TEXT,
            waktu_scrape TEXT
        )
    """)

    try:
        cur.execute("ALTER TABLE users ADD COLUMN password_plain TEXT")
    except:
        pass

    try:
        cur.execute("ALTER TABLE hasil_scraping ADD COLUMN username TEXT")
    except:
        pass

    try:
        cur.execute("ALTER TABLE riwayat_session ADD COLUMN username TEXT")
    except:
        pass

    # Buat admin default
    cur.execute("SELECT COUNT(*) FROM users WHERE username = 'admin'")
    if cur.fetchone()[0] == 0:
        cur.execute("""
            INSERT INTO users (username, password, password_plain, nama, role, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, ("admin", hash_pw("klhk2025"), "klhk2025", "Administrator", "admin",
              datetime.now().strftime("%Y-%m-%d %H:%M:?")))
        print("Akun admin default dibuat (admin / klhk2025)")

    conn.commit()
    cur.close()
    conn.close()
    print("Database SQLite siap!")

init_db()

# =============================================
# MODELS
# =============================================
class LoginRequest(BaseModel):
    username: str
    password: str

class ChangePasswordRequest(BaseModel):
    username: str
    password_lama: str
    password_baru: str

class TambahUserRequest(BaseModel):
    username: str
    password: str
    nama: str
    role: str = "user"

class ResetPasswordRequest(BaseModel):
    username: str
    password_baru: str

class ScrapeRequest(BaseModel):
    keyword: str
    platforms: List[str]
    max_pages: int = 3
    min_price: Optional[int] = 0
    max_price: Optional[int] = 999999999
    sort_by: Optional[str] = "relevance"
    username: Optional[str] = None   # ← tambahan

class ScrapeResultsRequest(BaseModel):
    session_id: str
    keyword: str
    username: str
    platforms: List[str]
    results: List[dict]

class ExportRequest(BaseModel):
    session_id: str
    keyword: str

# =============================================
# HEALTH CHECK
# =============================================
@app.get("/health")
def health_check():
    try:
        conn = get_conn()
        conn.close()
        return {"status": "ok", "app": "SiPantau", "versi": "1.0.0", "db": "PostgreSQL"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {str(e)}")

# =============================================
# AUTH
# =============================================
@app.post("/api/auth/login")
def login(req: LoginRequest, request: Request):
    ip = request.client.host if request.client else "unknown"
    check_rate_limit(ip)
    validate_input(req.username, "Username", max_length=50)
    if len(req.password) < 1 or len(req.password) > 100:
        raise HTTPException(status_code=400, detail="Password tidak valid")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (req.username,))
    user = cur.fetchone()
    cur.close(); conn.close()
    if not user or user["password"] != hash_pw(req.password):
        raise HTTPException(status_code=401, detail="Username atau password salah")
    return {"success": True, "user": {"username": user["username"], "nama": user["nama"], "role": user["role"]}}

@app.post("/api/auth/ganti-password")
def ganti_password(req: ChangePasswordRequest):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (req.username,))
    user = cur.fetchone()
    if not user or user["password"] != hash_pw(req.password_lama):
        cur.close(); conn.close()
        raise HTTPException(status_code=401, detail="Password lama salah")
    cur.execute("UPDATE users SET password = ?, password_plain = ? WHERE username = ?",
                (hash_pw(req.password_baru), req.password_baru, req.username))
    conn.commit(); cur.close(); conn.close()
    return {"success": True, "message": "Password berhasil diubah"}

# =============================================
# USER MANAGEMENT
# =============================================
@app.get("/api/users")
def get_users():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, username, password_plain, nama, role, created_at FROM users ORDER BY id")
    rows = cur.fetchall(); cur.close(); conn.close()
    return {"users": [dict(r) for r in rows]}

@app.post("/api/users")
def tambah_user(req: TambahUserRequest):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO users (username, password, password_plain, nama, role, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (req.username, hash_pw(req.password), req.password, req.nama, req.role,
              datetime.now().strftime("%Y-%m-%d %H:%M:?")))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.rollback(); cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Username sudah digunakan")
    cur.close(); conn.close()
    return {"success": True, "message": f"User '{req.username}' berhasil ditambahkan"}

@app.delete("/api/users/{username}")
def hapus_user(username: str):
    if username == "admin":
        raise HTTPException(status_code=400, detail="Akun admin tidak bisa dihapus")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE username = ?", (username,))
    if cur.rowcount == 0:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    conn.commit(); cur.close(); conn.close()
    return {"success": True, "message": f"User '{username}' berhasil dihapus"}

@app.post("/api/users/reset-password")
def reset_password_user(req: ResetPasswordRequest):
    """Admin reset password user tanpa perlu password lama"""
    if len(req.password_baru) < 6:
        raise HTTPException(status_code=400, detail="Password minimal 6 karakter")
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users WHERE username = ?", (req.username,))
    if cur.fetchone()[0] == 0:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    cur.execute("UPDATE users SET password = ?, password_plain = ? WHERE username = ?",
                (hash_pw(req.password_baru), req.password_baru, req.username))
    conn.commit(); cur.close(); conn.close()
    return {"success": True, "message": f"Password '{req.username}' berhasil direset"}

# =============================================
# SCRAPING — per user
# =============================================
@app.post("/api/scrape")
async def scrape(req: ScrapeRequest):
    session_id = datetime.now().strftime("%Y%m%d_%H%M?")
    username   = req.username or "unknown"
    all_results = []

    for platform in req.platforms:
        try:
            results = generate_placeholder(platform, req.keyword, req.max_pages)
            waktu   = datetime.now().strftime("%Y-%m-%d %H:%M:?")
            for r in results:
                r["waktu_scrape"] = waktu
                r["session_id"]   = session_id
            all_results.extend(results)
        except Exception as e:
            print(f"Error scraping {platform}: {e}")

    save_to_db(all_results, session_id, req.keyword, req.platforms, username)
    filename = export_to_excel_file(all_results, req.keyword, session_id)

    return {
        "session_id": session_id,
        "keyword":    req.keyword,
        "total":      len(all_results),
        "results":    all_results,
        "file_excel": filename
    }

@app.post("/api/scrape/results")
def receive_scrape_results(req: ScrapeResultsRequest):
    """
    Dipanggil oleh local agent setelah scraping selesai.
    Simpan hasil ke DB dan buat file Excel.
    """
    save_to_db(req.results, req.session_id, req.keyword, req.platforms, req.username)
    filename = export_to_excel_file(req.results, req.keyword, req.session_id)
    return {"success": True, "message": f"{len(req.results)} data disimpan", "file_excel": filename}

# =============================================
# EXPORT
# =============================================
@app.post("/api/export")
def export_excel(req: ExportRequest):
    conn = get_conn()
    df = pd.read_sql("SELECT * FROM hasil_scraping WHERE session_id = ?", conn, params=(req.session_id,))
    conn.close()
    if df.empty:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    filename = export_to_excel_file(df.to_dict("records"), req.keyword, req.session_id)
    return FileResponse(path=f"exports/{filename}", filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/export/download/{filename}")
def download_excel(filename: str):
    filepath = f"exports/{filename}"
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    return FileResponse(path=filepath, filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =============================================
# RIWAYAT — filter per user
# =============================================
@app.get("/api/riwayat")
def get_riwayat(username: str = ""):
    conn = get_conn()
    cur = conn.cursor()
    if username:
        cur.execute("SELECT * FROM riwayat_session WHERE username = ? ORDER BY id DESC LIMIT 50", (username,))
    else:
        cur.execute("SELECT * FROM riwayat_session ORDER BY id DESC LIMIT 50")
    rows = cur.fetchall(); cur.close(); conn.close()
    return {"riwayat": [dict(r) for r in rows]}

# =============================================
# STATS — per user
# =============================================
@app.get("/api/stats")
def get_stats(username: str = ""):
    conn = get_conn(); cur = conn.cursor()
    if username:
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE username = ?", (username,))
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE username = ? AND LOWER(platform) = 'tokopedia'", (username,))
        tokped = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE username = ? AND LOWER(platform) = 'shopee'", (username,))
        shopee = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE username = ? AND LOWER(platform) = 'lazada'", (username,))
        lazada = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM riwayat_session WHERE username = ?", (username,))
        ekspor = cur.fetchone()[0]
    else:
        cur.execute("SELECT COUNT(*) FROM hasil_scraping"); total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE LOWER(platform) = 'tokopedia'"); tokped = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE LOWER(platform) = 'shopee'");    shopee = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM hasil_scraping WHERE LOWER(platform) = 'lazada'");    lazada = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM riwayat_session"); ekspor = cur.fetchone()[0]
    cur.close(); conn.close()
    return {"total": total, "tokopedia": tokped, "shopee": shopee, "lazada": lazada, "ekspor": ekspor}

# =============================================
# HELPERS
# =============================================
def save_to_db(results, session_id, keyword, platforms, username="unknown"):
    conn = get_conn(); cur = conn.cursor()
    for r in results:
        cur.execute("""
            INSERT INTO hasil_scraping
            (session_id, username, keyword, nama_produk, harga, platform, rating, terjual, url_produk, gambar_url, waktu_scrape)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (session_id, username, keyword,
              r.get("nama_produk",""), r.get("harga",0),
              r.get("platform",""),   r.get("rating",0),
              r.get("terjual",""),    r.get("url_produk",""),
              r.get("gambar_url",""), r.get("waktu_scrape","")))
    cur.execute("""
        INSERT INTO riwayat_session (session_id, username, keyword, platforms, jumlah_data, status, waktu)
        VALUES (?,?,?,?,?,?,?)
    """, (session_id, username, keyword, ", ".join(platforms), len(results),
          "Selesai", datetime.now().strftime("%Y-%m-%d %H:%M:?")))
    conn.commit(); cur.close(); conn.close()


def export_to_excel_file(results, keyword, session_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    tanggal       = datetime.now().strftime("%Y-%m-%d")
    keyword_clean = keyword.replace(" ","_").replace("/","-")
    filename      = f"hasil_scraping_{keyword_clean}_{tanggal}_{session_id}.xlsx"
    filepath      = f"exports/{filename}"

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Data Scraping"

    ws.merge_cells("A1:I1")
    ws["A1"] = "KEMENTERIAN LINGKUNGAN HIDUP DAN KEHUTANAN REPUBLIK INDONESIA"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1B4332")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = f"SiPantau — Hasil Scraping: '{keyword}' | Tanggal: {tanggal}"
    ws["A2"].font = Font(bold=True, size=11, color="1B4332")
    ws["A2"].fill = PatternFill("solid", fgColor="D8F3DC")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    headers = ["No","Nama Produk","Harga (Rp)","Platform","Rating","Terjual","URL Produk","Waktu Scrape"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="2D6A4F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    for i, r in enumerate(results):
        row  = 5 + i
        is_expensive = r.get("harga", 0) >= 1000000
        if is_expensive:
            fill = PatternFill("solid", fgColor="FFCCCC") # Merah pucat
        else:
            fill = PatternFill("solid", fgColor="F0F7F4" if i%2==0 else "FFFFFF")
        data = [i+1, r.get("nama_produk",""), r.get("harga",0), r.get("platform",""),
                r.get("rating",0), r.get("terjual",""), r.get("url_produk",""), r.get("waktu_scrape","")]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill; cell.font = Font(size=9)
            if col == 3: cell.number_format = "#,##0"; cell.alignment = Alignment(horizontal="right")
            elif col == 5: cell.number_format = "0.0"

    col_widths = [6,45,18,15,10,12,50,22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    wb.save(filepath)
    return filename


def generate_placeholder(platform, keyword, max_pages):
    import random
    produk = [f"Kayu Jati {keyword}", f"Bambu {keyword}", f"Rotan {keyword}",
              f"Madu Hutan {keyword}", f"Gaharu {keyword}", f"Kayu Sengon {keyword}"]
    return [{
        "nama_produk": random.choice(produk) + f" #{random.randint(1,99)}",
        "harga":       random.randint(10000, 500000),
        "platform":    platform.capitalize(),
        "rating":      round(random.uniform(3.0, 5.0), 1),
        "terjual":     f"{random.randint(1,500)}rb+",
        "url_produk":  f"https://{platform}.co.id/produk/{keyword.replace(' ','-')}",
        "gambar_url":  ""
    } for _ in range(max_pages * 10)]


if __name__ == "__main__":
    import uvicorn
    print("SiPantau Backend (SQLite) berjalan di http://localhost:8000")
    uvicorn.run("main_sqlite:app", host="0.0.0.0", port=8000, reload=True)
